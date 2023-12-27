import csv
import os
import subprocess
import openpyxl
import customtkinter as ctk
from PIL import Image
from exchangelib import DELEGATE, Account, Credentials, Configuration, Message, FileAttachment
from datetime import date
import configparser
import shutil

#todo erstellen Сonfigvariable, um schnell mit dem Configparsermodul die Dateien zu ändern
config = configparser.ConfigParser()
config.read("configuration.ini", encoding='utf-8')

#todo variable mit laufendem Datum
date_today = date.today().strftime("%d.%m.%Y")
ctk.set_appearance_mode("Dark")

class App(ctk.CTk):

    def __init__(self):
        super().__init__()

        #todo Mainwindow einstellungen
        self.title("Argen-Digitale Inventur")                    #name
        self.screen_width = self.winfo_screenwidth()             #auflösung von Bildschirm
        self.screen_height = self.winfo_screenheight()           #auflösung von Bildschirm
        width = int(config['auflösung']['width'])                #Breit von window
        height = int(config['auflösung']['height'])              #Height von window
        x = (self.screen_width / 2) - (width / 2)                #Mit diesem Befehl platzieren wir das Fenster genau in die Mitte
        y = (self.screen_height / 2) - (height / 2)
        self.geometry(f"{width}x{height}+{int(x)}+{int(y)}")
        self.grid_columnconfigure(0, weight=1)                   #Wir teilen das Fenster in mehrere Teile auf, um die Elemente auf dem Bildschirm bequem anzuordnen und die Lesbarkeit zu verbessern.
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=1)

        #todo image erstellen
        img_dir = config['directory']['image_directory']         # Der Ordner, in dem sich alle Bilder befinden --- \\srv-file\Allgemein\Inventur_2023\images\
        self.iconbitmap(img_dir + "output.ico")
        self.image_mitarbeiter = ctk.CTkImage(Image.open(img_dir + "arbeiter.png"), size=(30, 30))
        self.image_exit = ctk.CTkImage(Image.open(img_dir + "exit.png"), size=(30, 30))
        self.image_option = ctk.CTkImage(Image.open(img_dir + "option.png"), size=(30, 30))
        self.image_back = ctk.CTkImage(Image.open(img_dir + "back.png"), size=(30, 30))
        self.image_change = ctk.CTkImage(Image.open(img_dir + "change.png"), size=(30, 30))
        self.image_send = ctk.CTkImage(Image.open(img_dir + "send.png"), size=(60, 60))
        self.image_in_process = ctk.CTkImage(Image.open(img_dir + "in_process.png"), size=(170, 170))
        self.image_disks = ctk.CTkImage(Image.open(img_dir + "disks.png"), size=(80, 80))
        self.image_fraese = ctk.CTkImage(Image.open(img_dir + "fraese1.png"), size=(80, 80))
        self.image_farbe = ctk.CTkImage(Image.open(img_dir + "farbe.png"), size=(80, 80))
        self.image_zubehoer = ctk.CTkImage(Image.open(img_dir + "zubehoer.png"), size=(80, 80))
        self.image_emax = ctk.CTkImage(Image.open(img_dir + "emax.png"), size=(120, 30))
        self.image_access = ctk.CTkImage(Image.open(img_dir + "access.png"), size=(100, 100))
        self.image_archiv = ctk.CTkImage(Image.open(img_dir + "archiv.png"), size=(200, 200))
        self.image_archiv_sw = ctk.CTkImage(Image.open(img_dir + "archiv_sw.png"), size=(200, 200))
        self.image_folder_mini = ctk.CTkImage(Image.open(img_dir + "folder.png"), size=(90, 90))
        self.image_plus = ctk.CTkImage(Image.open(img_dir + "plus.png"), size=(200, 200))
        self.image_plus_sw = ctk.CTkImage(Image.open(img_dir + "plus_sw.png"), size=(200, 200))
        self.image_dir = ctk.CTkImage(Image.open(img_dir + "directory.png"), size=(35, 35))
        self.image_email = ctk.CTkImage(Image.open(img_dir + "email.png"), size=(35, 35))
        self.image_sending = ctk.CTkImage(Image.open(img_dir + "sending.png"), size=(100, 100))
        self.image_argen_logo = ctk.CTkImage(dark_image=Image.open(img_dir + "ArgenLogo_Weiss.png"),
                                             light_image=Image.open(img_dir + "ArgenLogo_Schwarz.png"),
                                             size=(300, 50))

        #todo option frame, der in oberen Teil platziert worden ist
        self.options_frame = ctk.CTkFrame(self)                                       #Die Erstellung eines Optionsframes
        self.options_frame.grid_columnconfigure((0, 1, 3, 4, 5, 7, 8), weight=0)      #Die Erstellung eines visuellen Rasters
        self.options_frame.grid_columnconfigure((2, 6), weight=1)
        self.options_frame.grid_columnconfigure(4, minsize=150)
        self.options_frame.grid(row=0, sticky="we", pady=(10, 0), padx=10)            #Frame platzieren

        self.back_button = ctk.CTkButton(self.options_frame, text="", image=self.image_back)                                    #Back button
        self.change_user = ctk.CTkButton(self.options_frame, text="", image=self.image_mitarbeiter, command=self.login_frame_func)   #Change User button
        options_button = ctk.CTkButton(self.options_frame, text="", image=self.image_option, command=self.option_function)      #Einstellungen button
        options_button.grid(row=0, column=0, padx=(10, 5), pady=9, sticky="w")
        exit = ctk.CTkButton(self.options_frame, text="", image=self.image_exit, command=self.quit)                             #Programm schließen
        exit.grid(row=0, column=8, sticky="e", padx=(5, 10))
        self.name_label = ctk.CTkLabel(self.options_frame, font=ctk.CTkFont("Calibri", size=24, weight="bold"))                 #Name Label in Option Frame
        self.neu_send = ctk.CTkButton(self.options_frame, font=ctk.CTkFont("Calibri", size=18, weight="bold"),                  #Der Knopf, der csv. datei neuspeichert
                                      text="Erneut speichern  ", image=self.image_change)
        self.bool_info_liste = False        #Diese Boolean wurde erstellt, um zu definieren, ob Infoframe (Frame mit Liste) geoffnet worden ist. Um beim "Change user" die Fehler verbergen

        #todo login frame
        self.login_frame = ctk.CTkFrame(self, fg_color="transparent")                   #Die Erstellung eines Loginframes
        self.login_frame.grid_columnconfigure(0, weight=1)                              #Die Erstellung eines visuellen Rasters
        self.login_frame.grid_rowconfigure(0, weight=1)
        self.login_frame.grid_rowconfigure(3, weight=1)
        self.login_frame.grid(row=1, column=0, sticky="nsew", pady=10, padx=10)         #Frame platzieren
        self.user_list = config['user']['user_liste'].split(':')                        #Wir lesen die Informationen aus der Datei configuration.ini --- user_liste = Mark Wiesner:Ina Buch:Max Gerwers
        for num, user in enumerate(self.user_list):                                     #Die Erstellung einer Schaltfläche mit dem Benutzernamen
            ctk.CTkButton(self.login_frame, text=user, hover_color=("gray70", "gray30"), height=150, width=400,
                          fg_color="transparent", font=ctk.CTkFont(size=39, weight="bold"), corner_radius=10,
                          command=lambda user=user: self.login_password(user)).grid(row=num + 1, column=0)
            self.login_frame.grid_rowconfigure(num + 1, weight=0)
        max_len = len(self.user_list) + 1
        self.login_frame.grid_rowconfigure(max_len, weight=1)
        self.image_label = ctk.CTkLabel(self.login_frame, text="", image=self.image_argen_logo, anchor="s")  #Die Erstellung eines Labels mit Argenlogo
        self.image_label.grid(row=max_len, column=0, sticky="s", pady=25)

        #todo main frame, in diesem Frame befinden die Knopfen "Archiv" und "Neue Inventur"("Laufende Inventur")
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")                    #Die Erstellung eines Mainframes
        self.main_frame.grid_columnconfigure((0, 1), weight=1)                          #Die Erstellung eines visuellen Rasters
        self.main_frame.grid_rowconfigure(0, weight=1)
        self.archiv_button = ctk.CTkButton(self.main_frame, text="Archiv", image=self.image_archiv, width=400,
                                           height=400, fg_color="#50514F", hover_color="gray50", corner_radius=15,
                                           font=ctk.CTkFont("Calibri", size=45, weight="bold"), compound="top",
                                           command=self.archivs_griding_function)       #Die Erstellung eines "Archiv" Knopfes
        self.archiv_button.grid(row=0, column=0)

        #todo archiv frame, in dem sich das Archiv mit allen Archivordnern befindet
        self.archiv_frame = ctk.CTkScrollableFrame(self, orientation="vertical", fg_color="transparent")  #Die Erstellung eines Archivframes mit Scrollbar
        self.archiv_frame.grid_rowconfigure(0, weight=0)                                                  #Die Erstellung eines visuellen Rasters
        self.archiv_frame.grid_rowconfigure(1, weight=1)
        self.archiv_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
        self.email_liste = config['email']['empfaengers'].split(':')                                      #empfaengers = s.orsouw@argen.de:s.fedkenhauer@argen.de

        #todo artikels frame, in dem sich alle aktuellen Inventurlisten befinden
        self.artikels_frame = ctk.CTkFrame(self, fg_color="transparent")                                  #Die Erstellung eines Artikelframes
        self.artikels_frame.grid_columnconfigure(0, weight=1)
        self.artikels_frame.grid_rowconfigure((0, 1, 2, 3, 4, 5, 6), weight=1)

        self.inventur_end = ctk.CTkButton(self.artikels_frame, width=500, height=100,                     #Die Erstellung eines "Inventur abschließen" Knopfes
                                          text="Inventur abschließen",
                                          font=ctk.CTkFont("Calibri", size=40, weight="bold"),
                                          compound="left", hover_color="#5FAD56",
                                          fg_color="gray20", corner_radius=10, command=self.invent_end_func)
        self.inventur_end.grid(row=6, column=0, pady=25)


    #todo: Die funktionen für Back Button
    def login_frame_func(self):     #Die Funktion, die den ursprünglichen Bildschirm aufruft.

        """"Wenn Boolean "self.bool_info_liste" ist "True", List enthalt "info_frame",
        wenn Boolean "self.bool_info_liste" ist "False" oder "Else", 
        List enthalt keine "info_frame", um Fehler zu vermeiden.
        
        Wenn "info_frame" nicht geoffnet wurden, 
        können wir nicht es löchen um "Loginframe" zu platzieren."""""

        if self.bool_info_liste:
            frame_list = [self.back_button, self.change_user, self.name_label,
                          self.neu_send, self.main_frame, self.archiv_frame,
                          self.artikels_frame, self.info_frame]
        else:
            frame_list = [self.back_button, self.change_user, self.name_label,
                          self.neu_send,self.main_frame, self.archiv_frame,
                          self.artikels_frame]
        for i in frame_list:                                                            #Löchen alle "Widgets" und "Frames", die in die "frame_list" sind
            i.grid_forget()
        self.login_frame.grid(row=1, column=0, sticky="nsew", pady=10, padx=10)         #Loginframe platzieren

    def back_to_run(self):

        """"Zurück zum Mainframe"""""

        self.artikels_frame.grid_forget()                                               #Artikelframe entfernen
        self.archiv_frame.grid_forget()                                                 #Archivframe entfernen
        self.main_frame.grid(row=1, column=0, sticky="nsew", pady=10, padx=10)          #Mainframe platzieren
        self.back_button.configure(command=self.login_frame_func)                       #Wir ändern die Funktion des Back Button: "login_frame_func"
    def back_archiv_func(self):

        """"Zurück zum Archiv"""""

        self.artikels_frame.grid_forget()                                               #Artikelframe entfernen
        self.archiv_frame.grid(row=1, sticky="nsew", pady=(10, 0), padx=10)             #Archivframe platzieren
        self.back_button.configure(command=self.back_to_run)                            #Ändern die Funktion des Back Button: "back_to_run"
        self.name_label.configure(text=self.user)                                       #Ändern label (in Optionsframe) um Benutzername zu zeigen
    def back_button_func_data(self, num):                                               #In dieser Funktion übergeben wir eine "num" Variable, die bestimmt, welches Fenster geöffnet wird, wenn die Zurück-Taste gedrückt wird.
        if num==0:                                                                      #Wenn num "0" ist, übernimmt der Knopf "back" die Funktion "back_archiv_func"
            self.back_button.configure(command=self.back_archiv_func)
        else:                                                                           #Wenn num "1" oder andere ist, übernimmt der Knopf "back" die Funktion "back_to_run"
            self.back_button.configure(command=self.back_to_run)
        self.neu_send.grid_forget()                                                     #Neusend button entfernen, weil dieser Knopf nur in Archivframe und in Optionspanelle platzieren seien muss
        for i in self.info_frame.winfo_children():                                      #Entfernen alle Widgets von "info_frame"
            i.grid_forget()
        self.info_frame.grid_forget()                                                   #Entfernen "info_frame"
        self.artikels_frame.grid(row=1, column=0, sticky="nsew", pady=10, padx=10)      #Artikelframe platzieren

    #todo: Die Hauptfunktionen
    def login_password(self, user):

        """"In dieser Funktion erstellen wir ein Fenster zur Passworteingabe.
        Dabei übergeben wir den Benutzernamen "user", der versucht, sich im System zu registrieren,
        um diesen Namen später auch im Optionsframe zu verwenden."""""

        self.user = user
        x = (self.screen_width / 2) - (500 / 2)                                         #Skalierung
        y = (self.screen_height / 2) - (280 / 2)

        password_frame = ctk.CTkToplevel(self.login_frame, fg_color="gray")             #Fenster erstellen
        password_frame.overrideredirect(True)                                           #Diese Zeile bedeutet, dass das Fenster keinen Statusbar haben wird
        password_frame.geometry(f"{500}x{280}+{int(x)}+{int(y)}")
        password_frame.grab_set()                                                       #Das Fenster wird über den anderen liegen.
        self.entry_feld_bool = False
        password_frame.grid_columnconfigure((0,1), weight=1)
        ctk.CTkLabel(password_frame, text=user, font=ctk.CTkFont("Calibri", size=33, weight="bold")).grid(row=0,
                                                                                column=0, columnspan=2, pady=(35,30))   #Username label
        password = ctk.CTkEntry(password_frame, width=390, height=40, placeholder_text="Password", show="*",            #Passwordfeld
                                font=ctk.CTkFont("Calibri", size=25, weight="bold"))
        password.grid(row=1, column=0, columnspan=2, pady=(0,35))


        def start_eingabe(e):

            """""Beim Drücken der "Eingabe"-Taste wird der Befehl aufgerufen und das Passwort überprüft."""""
            start()

        def start():

            """"Wir überprüfen, ob das Passwort korrekt ist."""""

            if config['password'][user] == password.get():                #ob das Passwort korrekt ist
                if user in config['admins']['admins'].split(":"):         #Wenn ein Mitarbeiter in der Liste der Administratoren steht, hat er Administratorrechte und Zugriff auf entsprechende Funktionen und Berechtigungen.
                    password_frame.destroy()                              #Fragefenster entfernen
                    self.admins_bool = True                               #Ist Admin
                    self.run(self.admins_bool)                            #Funktion "run" starten
                else:
                    self.admins_bool = False                              #Ist nicht Admin
                    password_frame.destroy()                              #Fragefenster entfernen
                    self.run(self.admins_bool)                            #Funktion "run" starten
            else:
                ctk.CTkLabel(password_frame, text="Password ist nicht korrekt", text_color="orange",   #Error mitteilung
                             font=ctk.CTkFont("Calibri", size=28, weight="bold")).grid(row=3, column=0, columnspan=2)

        password.bind('<Return>', start_eingabe)                          #Wir weisen der Schaltfläche eine Funktion zu.
        password_input_btn = ctk.CTkButton(password_frame, text="Ok", font=ctk.CTkFont("Calibri", size=25, weight="bold"), command=start).grid(row=2, column=0, pady=(0,15))         #ok ("start" funktions)
        exit = ctk.CTkButton(password_frame, text="Exit", command=password_frame.destroy, font=ctk.CTkFont("Calibri", size=25, weight="bold")).grid(row=2, column=1, pady=(0,15))    #exit

    def run(self, admins_bool):

        """"Der Hauptbildschirm, auf dem die Schaltflächen "Archiv" und "Start" (Fortsetzen) der Inventarisierung angezeigt werden."""""

        if admins_bool:                                                                       #Wenn der Mitarbeiter ein Administrator ist, wird die Start-Schaltfläche für die Inventarisierung freigeschaltet.
            self.archiv_button.configure(state="normal", image=self.image_archiv)
        else:                                                                                 #Wenn es sich um einen normalen Benutzer handelt, wird die Archiv-Schaltfläche für die Inventarisierung deaktiviert sein.
            self.archiv_button.configure(state="disabled", image=self.image_archiv_sw)

        self.login_frame.grid_forget()                                                        #Loginframe entfernen
        self.main_frame.grid(row=1, column=0, sticky="nsew", pady=10, padx=10)                #Mainframe platzieren
        if len(os.listdir(config['directory']['neu_inventur_dir'])) == 1:    #Wenn sich in dem Ordner, in dem die neue Inventarisierung erstellt wird, nur eine Datei (Archivordner) befindet,
            self.invent_button = ctk.CTkButton(self.main_frame, text="Neue Inventur", image=self.image_plus, width=400,  #wird die Inventur-Schaltfläche als "Neu Inventur" angezeigt. ----------------------------------------------------->
                                               height=400, compound="top", corner_radius=15, hover_color="gray50",
                                               fg_color="#50514F", font=ctk.CTkFont("Calibri", size=45, weight="bold"),
                                               command=self.neu_inventur)
            self.invent_button.grid(row=0, column=1)                                            #Knopf platzieren
            if admins_bool:                                                                     #Wenn der Mitarbeiter ein Administrator ist,
                self.invent_button.configure(state="normal")                                    #wird die Archiv-Schaltfläche für die Inventarisierung aktivieren.
            else:                                                                               #Ein normaler Benutzer kann nur Änderungen an der aktuellen Inventarisierung vornehmen, die vom Administrator gestartet wurde.
                self.invent_button.configure(state="disabled", image=self.image_plus_sw)
        else:
            self.invent_button = ctk.CTkButton(self.main_frame, text=f"Inventur\n{date_today}",                          #-------------------------------------------> Andernfalls wird die Schaltfläche als "Aktuelle Inventur" angezeigt.
                                               image=self.image_in_process, width=400,
                                               height=400,
                                               fg_color="#50514F",
                                               font=ctk.CTkFont("Calibri", size=45, weight="bold"),
                                               compound="top", corner_radius=15, hover_color="gray50",
                                               command=lambda file_dir=config['directory'][
                                               'neu_inventur_dir']: self.artikel_frame_griding(file_dir, 1))
            self.invent_button.grid(row=0, column=1)                                            #Knopf platzieren

        self.back_button.grid(row=0, column=1, sticky="w")                                      #Back button platzieren
        self.change_user.grid(row=0, column=7, sticky="e")                                      #Change-User button platzieren
        self.back_button.configure(command=self.login_frame_func)                               #übernimmt der Knopf "back" die Funktion "login_frame_func"
        self.name_label.configure(text=self.user)                                               #In Optioinsframe wird User-Name angezeigt
        self.name_label.grid(row=0, column=4)                                                   #User-Name Label platzieren
    def neu_inventur(self):

        """"In dieser Funktion erstellen wir ein Fenster mit der Frage, ob wir wirklich eine neue Inventarisierung starten möchten."""""

        x = (self.screen_width / 2) - (500 / 2)                                                 #Skalierung
        y = (self.screen_height / 2) - (150 / 2)
        self.dialog_fenster = ctk.CTkToplevel(self.main_frame, fg_color="gray")                 #Fenster erstellen
        self.dialog_fenster.overrideredirect(True)                                              #Diese Zeile bedeutet, dass das Fenster keinen Statusbar haben wird
        self.dialog_fenster.geometry(f"{500}x{150}+{int(x)}+{int(y)}")
        self.dialog_fenster.grab_set()                                                          #Das Fenster wird über den anderen liegen.
        self.dialog_fenster.grid_rowconfigure((0,1), weight=1)                                  #Die Erstellung eines visuellen Rasters
        self.dialog_fenster.grid_columnconfigure((0,1), weight=1)
        ctk.CTkLabel(self.dialog_fenster, text="Sind Sie sicher?", font=ctk.CTkFont("Calibri", weight="bold", size=39)).grid(row=0, column=0, columnspan=2)      #Die Erstellung "Sind Sie sicher?" Label
        ctk.CTkButton(self.dialog_fenster, text="Nein", font=ctk.CTkFont("Calibri", weight="bold", size=29),                                                     #Exit
                      command=self.dialog_fenster.destroy).grid(row=1, column=1)
        ctk.CTkButton(self.dialog_fenster, text="Ja", font=ctk.CTkFont("Calibri", weight="bold", size=29), command=self.datei_erstellen).grid(row=1, column=0)   # "Ja" weiter zum "datei_erstellen"

    def datei_erstellen(self):

        """"Wir erstellen einen Pool von Dateien für zukünftige Inventarisierungszwecke."""""

        self.dialog_fenster.destroy()                                                                   #Dialog Fenster entfernen
        self.default_listen = os.listdir(config['directory']['default_liste_dir'])                      #Wir erhalten eine Liste der Dateien, die im Ordner "Default" gespeichert sind, um diese leeren Listen für die Inventarisierung zu verwenden.
        for i in self.default_listen:                                                                   #Wir öffnen jede Excel-Datei,....
            book = openpyxl.open(config['directory']['default_liste_dir'] + i)                          #...die sich im Ordner "Default" befindet,...
            book.save(config['directory']['neu_inventur_dir'] + "Inv_" + date_today + "_" + i)          #...und benennen sie in das Format "Inv_(datum)_(typ).xlsx" um...
            book.close()                                                                                #...anschließend speichern wir sie im Ordner der neuen Inventarisierung, und schließen excell Buch.
        self.invent_button.configure(command=lambda file_dir = config['directory']['neu_inventur_dir']: self.artikel_frame_griding(file_dir, 1),
                                     text=f"Inventur\n{date_today}", image=self.image_in_process)       #Wir ändern die zuvor zugewiesene Funktion für die Schaltfläche "Invent" und aktualisieren auch das Bild und den Namen.
        self.artikel_frame_griding(config['directory']['neu_inventur_dir'], 1)
    def artikel_frame_griding(self, directory, num):

        """"Wir erstellen die Knöpfe, um die Inventarlisten-Dateien zu öffnen."""""

        """"Hier bestimmen wir, auf welche Weise der Frame geöffnet wurde, 
        um beim Klicken der Zurück-Schaltfläche je nach Fall entweder zum 
        Bildschirm mit den Archivordnern zurückzukehren oder zum 
        Hauptbildschirm mit der Auswahl zwischen Inventarisierung 
        und Archiv."""""

        if num == 0:
            self.back_button.configure(command=self.back_archiv_func)
            self.inventur_end.configure(state="disabled")
            self.name_label.configure(text=fr"{self.user} - {directory[-15:-1]}")
        else:
            self.back_button.configure(command=self.back_to_run)
            self.inventur_end.configure(state="normal")
        self.main_frame.grid_forget()
        self.archiv_frame.grid_forget()

        self.artikels_frame.grid(row=1, column=0, sticky="nsew", pady=10, padx=10)                          #Artikelframe platzieren
        for files in os.listdir(directory):                                                                 #Definieren für jede Artikeltypliste bestimmte Excell file(directory)
            if config['artikel_typ']['discs'] in files:
                self.disks_file = directory + files
            elif config['artikel_typ']['fräser'] in files:
                self.fras_file = directory + files
            elif config['artikel_typ']['EMAX'] in files:
                self.emax_file = directory + files
            elif config['artikel_typ']['Drücker_Zub'] in files:
                self.druecker_file = directory + files
            elif config['artikel_typ']['Zirkon_farbe'] in files:
                self.farbe_file = directory + files

        self.files_liste = [self.disks_file, self.fras_file, self.emax_file, self.druecker_file, self.farbe_file]       #Die Erstellung eine Liste mit Directory von alle Dateityp

        height_art_button = int(config['auflösung']['width_artikel_button'])                                            #In der ini-Datei können wir jederzeit die Höhe der Schaltflächen an die Bildschirmauflösung anpassen.

        #Die Knöpfe für Artikellisteframe erstellen
        self.disks = ctk.CTkButton(self.artikels_frame, width=700, image=self.image_disks, height=height_art_button,
                                 text="         " + config['artikel_typ']['discs'],
                                 font=ctk.CTkFont("Calibri", size=45, weight="bold"), compound="left",
                                 hover_color="#F2C14E", anchor="w",
                                 fg_color="gray20", corner_radius=10, command=lambda file=self.disks_file: self.data_frame_function(file, config['artikel_typ']['discs'], "#F2C14E", num)).grid(row=1, column=0)
        self.fras = ctk.CTkButton(self.artikels_frame, width=700, height=height_art_button, image=self.image_fraese,
                                 text="         " + config['artikel_typ']['fräser'], font=ctk.CTkFont("Calibri", size=45, weight="bold"),
                                 compound="left", hover_color="#F78154", anchor="w",
                                 fg_color="gray20", corner_radius=10, command=lambda file=self.fras_file: self.data_frame_function(file, config['artikel_typ']['fräser'], "#F78154", num)).grid(row=2, column=0)
        self.emax = ctk.CTkButton(self.artikels_frame, width=700, height=height_art_button, image=self.image_emax,
                                 text="     " + config['artikel_typ']['EMAX'], font=ctk.CTkFont("Calibri", size=45, weight="bold"),
                                 compound="left", hover_color="#B4436C", anchor="w",
                                 fg_color="gray20", corner_radius=10, command=lambda file=self.emax_file: self.data_frame_function(file, config['artikel_typ']['EMAX'], "#B4436C", num)).grid(row=3, column=0)
        self.d_zubeh = ctk.CTkButton(self.artikels_frame, width=700, height=height_art_button, image=self.image_zubehoer,
                                 text="         " +config['artikel_typ']['Drücker_Zub'],
                                 font=ctk.CTkFont("Calibri", size=45, weight="bold"), anchor="w",
                                 compound="left", hover_color="#4D9078",
                                 fg_color="gray20", corner_radius=10, command=lambda file=self.druecker_file: self.data_frame_function(file, config['artikel_typ']['Drücker_Zub'], "#4D9078", num)).grid(row=4, column=0)
        self.zirkon_farb = ctk.CTkButton(self.artikels_frame, width=700, height=height_art_button, image=self.image_farbe,
                                 text="         " +config['artikel_typ']['Zirkon_farbe'],
                                 font=ctk.CTkFont("Calibri", size=45, weight="bold"), anchor="w",
                                 compound="left", hover_color="#3d91cc",
                                 fg_color="gray20", corner_radius=10, command=lambda file=self.farbe_file: self.data_frame_function(file, config['artikel_typ']['Zirkon_farbe'], "#3d91cc", num)).grid(row=5, column=0)

    def data_frame_function(self, file, name, farbe, num):

        """"Eine der wichtigsten Funktionen des Programms liest Daten von den erforderlichen Excel-Blättern aus. 
        Basierend auf den Daten in diesen Blättern werden Labels mit Artikeln und 
        Produktnamen auf dem Bildschirm angezeigt, sowie Eingabefelder für die Menge der Produkte. 
        Bei jeder Änderung der Informationen in den Eingabefeldern wird die entsprechende 
        Excel-Datei automatisch gespeichert.
        
        In dieser Funktion werden auch die folgenden Argumente übergeben: "file", "name", "farbe" und "num". 
        
        "File" ist der Pfad zur Datei, 
        "name" ist der Dateiname, 
        "farbe" ist die Farbe, die jedem Dateityp zur erleichterten visuellen Unterscheidung zugewiesen wird, 
        "Num" ist ein funktionaler Argument, das entweder den Wert 0 oder 1 haben kann."""""

        if num == 0:
            folder_path = os.path.dirname(file)                                                                         #Der Dateiname, der sich in diesem Verzeichnis befindet.
            end_dir = folder_path + "\\" + "upload" + "\\" + name                                                       #Der Ordner, in dem sich die zu ersetzende Datei befindet.
            file_ersetz = os.listdir(end_dir)[0]                                                                        #Wir weisen der Variable den Namen der endgültigen Datei zu.
            file_ersetz_dir = end_dir + "\\" + file_ersetz                                                              #Die endgültige Datei mit dem Verzeichnis, in dem der Ersatz stattfinden soll.

            self.neu_send.grid(row=0, column=6)                                                                         #Platzieren die Schaltfläche zum Überschreiben der Datei.
            self.neu_send.configure(command = lambda csv_file = file_ersetz_dir : self.neu_writing(file, csv_file))     #Wir ändern die Eigenschaften der Schaltfläche und weisen ihr die Funktion (self.neu_writing) als Befehl zu.
            self.back_button.configure(command=lambda : self.back_button_func_data(0))                                  #Weisen wir Funktion für "back_button" zu.
        else:
            self.back_button.configure(command=lambda : self.back_button_func_data(1))                                  #Weisen wir Funktion für "back_button" zu.

        self.info_frame = ctk.CTkScrollableFrame(self, label_text_color="black", label_font=ctk.CTkFont("Calibri", size=22))    #Wir erstellen einen Rahmen (Frame), in dem wir unsere Liste mit Eingabefeldern für die Menge platzieren werden.
        self.bool_info_liste = True                                                                     #Das bedeutet, dass der Frame erstellt wurde und beim nächsten Aufruf der "change_user" Funktion wird er in der Liste zum Entfernen aus dem Fenster stehen.

        book = openpyxl.open(file)                                                                                      #Wir öffnen eine Excel-Datei mit der benötigten Liste.
        sheet = book.active                                                                                             #Wir wählen das aktive Arbeitsblatt in der Excel-Datei aus.
        self.artikels_frame.grid_forget()                                                                               #Entfernen den Frame.
        self.info_frame.configure(label_text=name, label_fg_color=farbe)                                                #Wir passen den Frame an und setzen den Dateinamen sowie die Farbe in seinem oberen Teil, um die visuelle Identifizierung zu erleichtern.
        self.info_frame.grid(row=1, column=0, sticky="nsew", pady=10, padx=10)                                          #Platzieren den Frame
        self.entry_feld_bool = True
        entry_list = []                                                                                                 #Wir erstellen eine entry_list, um in Zukunft die Eingabefelder zu füllen und auf sie über ihren Index in dieser Liste zugreifen zu können.
        for num in range(1, sheet.max_row + 1):                                                                         #Wir starten eine Schleife für jede Zeile in der Excel-Datei.
            ctk.CTkLabel(self.info_frame, font=ctk.CTkFont("Calibri", size=24),                                         #Wir erstellen ein Label, in das wir die Informationen aus der ersten und zweiten Spalte der Excel-Datei einfügen(bzw. den Artikel und den Namen.)
                         text=f"{sheet.cell(num, column=1).value}    {sheet.cell(num, column=2).value}").grid(
                         row=num - 1, column=0, sticky="w", pady=3, padx=15)
            if name == "Drucker Zubehör":                                                                               #Die Datei "Drücker Zubehör" enthält Artikel,...
                ctk.CTkLabel(self.info_frame, font=ctk.CTkFont("Calibri", size=24),                                     #...die nicht pro Stück, sondern in Millilitern oder Packungen gezählt werden. --->
                         text=sheet.cell(num, column=4).value).grid(row=num - 1, column=2, sticky="w", pady=3, padx=15) #...Daher verwenden wir als Maßeinheit für diese Produkte die Daten aus der vierten Spalte der Excel-Datei.
            else:                                                                                                       #In allen anderen Fällen verwenden wir einfach "stk." als Einheit überall
                ctk.CTkLabel(self.info_frame, font=ctk.CTkFont("Calibri", size=24),
                             text="stk.").grid(row=num - 1, column=2, sticky="w", pady=3, padx=15)

            self.entry = ctk.CTkEntry(self.info_frame, width=100, font=ctk.CTkFont("Calibri", size=24), height=40, corner_radius=8)      #Wir erstellen ein Eingabefeld.
            self.entry.grid(row=num - 1, column=1, pady=3, padx=15)                                                                      #Eingabefeld platzieren
            entry_list.append(self.entry)                                                                                                #Wir fügen der entry_list Eingabefelder hinzu.
            if sheet.cell(num, column=3).value != None:                                                                 #Wenn eine Zelle in der Excel-Datei nicht leer ist,...
                entry_list[num-1].insert("0", sheet.cell(num, column=3).value)                                          #...fügen wir die Informationen in das entsprechende Eingabefeld ein.
            else:                                                                                                       #Andernfalls überspringen wir diese Aktion.
                pass

            def get_cursor(event, num=num):                                                                             #Eine Funktion, die nach jeder Aktion in den Eingabefeldern Informationen in die Excell-Datei einfügt
                if self.entry_feld_bool:
                    if entry_list[num-1].get() == '':                                                                       #Wenn das Eingabefeld leer ist,...
                        sheet.cell(num, column=3).value = None                                                              #...löschen wir die Zelle in Excell
                    else:
                        sheet.cell(num, column=3).value = int(entry_list[num-1].get())                                      #Im anderen Fall fügen Sie die Daten aus dem Eingabefeld in eine Excell-Datei ein
                        for col in range(5,15):
                            if sheet.cell(num, column=col).value == None:                                                   #Wir fügen auch Informationen darüber hinzu, wer die Änderungen in dieser Zeile vorgenommen hat und welche Menge eingegeben wurde.
                                sheet.cell(num, column=col).value = self.user + f" -> ({int(entry_list[num-1].get())})"
                                break
                    book.save(file)                                                                                       #Speichern
                else:
                    pass
            self.entry.bind("<KeyRelease>", get_cursor)                                                                      #Ein Ereignis zuweisen, wird die Funktion "get_cursor" jedes Mal aufgerufen, wenn eine Schaltfläche im Eingabefeld gedrückt wird

    def neu_writing(self, file, csv_file):

        """"Diese Funktion erstellt ein Fenster mit einer Frage, 
        deren Antworten entweder "Ja" oder "Nein" sind. Wenn die Antwort "Ja" lautet, 
        wird die Funktion "csv_ersetz" aufgerufen. Wenn die Antwort "Nein" lautet, 
        wird das Fenster geschlossen.
        
        In dieser Funktion übergeben wir auch zwei Argumente: "file" und "csv_file". 
        "file" ist der Name des Excel-Blattes, von dem neue Informationen gelesen werden sollen 
        (zum Beispiel: Disks, Emax, Tauchfarbe). "csv_file" ist der Pfad zur CSV-Datei, 
        die überschrieben werden soll."""""

        x = (self.screen_width / 2) - (500 / 2)                                                                         #Fenster platzieren
        y = (self.screen_height / 2) - (150 / 2)

        self.csv_dialog_fenster = ctk.CTkToplevel(self.artikels_frame, fg_color="gray")                                 #Fenster erstellen
        self.csv_dialog_fenster.overrideredirect(True)                                                                  #Diese Zeile bedeutet, dass das Fenster keinen Statusbar haben wird
        self.csv_dialog_fenster.geometry(f"{500}x{150}+{int(x)}+{int(y)}")                                              #Skallieren
        self.csv_dialog_fenster.grab_set()                                                                              #Das Fenster wird über den anderen liegen.
        self.csv_dialog_fenster.grid_rowconfigure((0, 1), weight=1)                                                     #Die Erstellung eines visuellen Rasters
        self.csv_dialog_fenster.grid_columnconfigure((0, 1), weight=1)                                                  #Die Erstellung eines visuellen Rasters
        ctk.CTkLabel(self.csv_dialog_fenster, text="Sind Sie sicher?",                                                  #Die Erstellung "Sind Sie sicher?" Label
                     font=ctk.CTkFont("Calibri", weight="bold", size=39)).grid(row=0, column=0, columnspan=2)
        ctk.CTkButton(self.csv_dialog_fenster, text="Nein", font=ctk.CTkFont("Calibri", weight="bold", size=29),        #Exit
                      command=self.csv_dialog_fenster.destroy).grid(row=1, column=1)
        ctk.CTkButton(self.csv_dialog_fenster, text="Ja", font=ctk.CTkFont("Calibri", weight="bold", size=29),          #Rufen wir Funktion "csv_ersetz" auf
                      command=lambda : self.csv_ersetz(file, csv_file)).grid(row=1, column=0)
    def csv_ersetz(self, file, csv_file):

        """"Diese Funktion aktualisiert die .csv-Datei, falls zuvor ein Fehler in der Inventur aufgetreten ist."""""

        self.csv_dialog_fenster.destroy()                                                                               #Fenster entfernen
        book = openpyxl.open(file)                                                                                      #Wir öffnen erneut die benötigte Excel-Datei, um die CSV-Datei zu überschreiben.
        sheet = book.active                                                                                             #Wir wählen das aktive Arbeitsblatt in der Excel-Datei aus.
        with open(csv_file, "w", newline='') as file_csv:                                                               #Wir öffnen die Datei, fügen weitere Informationen hinzu oder ändern sie und speichern die CSV-Datei in einem bestimmten Ordner (csv_file).
            writer = csv.writer(file_csv, delimiter=";")
            for rows in range(1, sheet.max_row + 1):
                value = [sheet.cell(row=rows, column=1).value, sheet.cell(row=rows, column=3).value]
                writer.writerow(value)
    def invent_end_func(self):

        x = (self.screen_width / 2) - (500 / 2)                                                                         #Fenster platzieren
        y = (self.screen_height / 2) - (250 / 2)

        self.dialog_fenster = ctk.CTkToplevel(self.artikels_frame, fg_color="gray")                                     #Fenster erstellen
        self.dialog_fenster.overrideredirect(True)                                                                      #Diese Zeile bedeutet, dass das Fenster keinen Statusbar haben wird
        self.dialog_fenster.geometry(f"{500}x{250}+{int(x)}+{int(y)}")                                                  #Skallieren
        self.dialog_fenster.grab_set()                                                                                  #Das Fenster wird über den anderen liegen.
        self.dialog_fenster.grid_columnconfigure((0, 1), weight=1)

        def daten_speichern_funktion():

            """"Die Hauptfunktion besteht darin, die Inventurdateien und CSV-Dateien zu speichern."""""

            self.archiv_invent = config['directory']['archiv_dir'] + "Inv_" + date_today                                #Weg zum Archivpfad
            try:                                                                                                        #bei dem Versuch...
                os.makedirs(self.archiv_invent)                                                                         #Wir erstellen Archivordner mit dem aktuellen Datum
                for files in self.files_liste:
                    shutil.move(files, self.archiv_invent)                                                              #Alle Dateien in den soeben erstellten Ordner übertragen
                self.artikels_frame.grid_forget()                                                                       #Artikel frame entfernen
                self.dialog_fenster.destroy()                                                                           #Dialog fenster enfernen
                self.finish_func()                                                                                      #Finish funktion aufrufen
                self.run(self.user)                                                                                     #"Run" funktions aufrufen, um "Mainframe" zu platzieren
            except FileExistsError:                                                                                     #Wenn ein Ordner mit diesem Namen bereits existiert, zeigt das Programm eine Fehlermeldung an
                self.ja.grid_forget()                                                                                   #Entfernen "ja" button
                self.nein.grid_forget()                                                                                 #Entfernen "nein" button
                self.label.configure(text="Fehler beim Speichern der Datei.\nEin Ordner mit diesem Namen\nexistiert bereits.",
                                     text_color="#F29E4A", font=ctk.CTkFont("Calibri",weight="bold", size=26))          #Erstellen den Text mit dem Fehler ausgeben
                ctk.CTkButton(self.dialog_fenster, text="Exit", font=ctk.CTkFont("Calibri",weight="bold", size=26),     #Exit button
                              command=self.dialog_fenster.destroy).grid(row=3, columnspan=2, pady=(0, 45))

            upload_dir = self.archiv_invent + "\\" + config['directory']['upload_file']                                 #Erstellen einen Ordrer, in dem alle .csv dateien gespeichert werden
            os.makedirs(upload_dir)
            list = [config['artikel_typ']['discs'], config['artikel_typ']['fräser'],
                    config['artikel_typ']['EMAX'], config['artikel_typ']['Drücker_Zub'],                                #Eine Liste mit verschiedenen Dateitypen erstellen
                    config['artikel_typ']['Zirkon_farbe']]
            for typ in list:
                upload_file = upload_dir + "\\" + typ                                                                   #Einen Ordner für jeden Dateityp erstellen
                os.makedirs(upload_file)

            archiv_list = os.listdir(self.archiv_invent)                                                                #Erstellen eine Liste von Dateinamen, die sich im aktuellen Ordner befinden
            archiv_list.remove(config['directory']['upload_file'])                                                      #Entfernen "archiv" Name
            for files in archiv_list:                                                                                   #Für jeden Dateityp wird ein bestimmter Endordner für die Speicherung der CSV-Datei definiert.
                address = self.archiv_invent + "\\" + files
                book = openpyxl.open(address)
                sheet = book.active
                for file_dirs in os.listdir(upload_dir):
                    if config['artikel_typ']['discs'] in files:
                        full_file = upload_dir + "\\" + config['artikel_typ']['discs']
                    if config['artikel_typ']['Drücker_Zub'] in files:
                        full_file = upload_dir + "\\" + config['artikel_typ']['Drücker_Zub']
                    if config['artikel_typ']['EMAX'] in files:
                        full_file = upload_dir + "\\" + config['artikel_typ']['EMAX']
                    if config['artikel_typ']['Zirkon_farbe'] in files:
                        full_file = upload_dir + "\\" + config['artikel_typ']['Zirkon_farbe']
                    if config['artikel_typ']['fräser'] in files:
                        full_file = upload_dir + "\\" + config['artikel_typ']['fräser']

                with open(full_file + "\\" + files[:14] + ".csv", "w", newline='') as file_csv:                         #Wir lesen die erforderlichen Daten aus der Excel-Datei und erstellen eine CSV-Datei im gewünschten Format.
                    writer = csv.writer(file_csv, delimiter=";")
                    for rows in range(1, sheet.max_row + 1):
                        value = [sheet.cell(row=rows, column=1).value, sheet.cell(row=rows, column=3).value]
                        writer.writerow(value)

        self.label = ctk.CTkLabel(self.dialog_fenster, text="Sind Sie sicher?\nNach Abschluss wird die aktuelle"        #Label mit Frage erstellen
                                                            "\nInventur in das Archiv verschoben.",
                                                            font=ctk.CTkFont("Calibri", weight="bold", size=30))
        self.label.grid(row=0, column=0, columnspan=2, pady=30)                                                         #Label platzieren
        self.ja = ctk.CTkButton(self.dialog_fenster, text="Ja", command=daten_speichern_funktion,                       #"Ja" button
                                font=ctk.CTkFont("Calibri", weight="bold", size=29))
        self.ja.grid(row=1, column=0, pady=15)                                                                          #Button platzieren
        self.nein = ctk.CTkButton(self.dialog_fenster, text="Nein", font=ctk.CTkFont("Calibri", weight="bold", size=29),#"Nein" button
                                                       command=self.dialog_fenster.destroy)
        self.nein.grid(row=1, column=1, pady=15)                                                                        #Button platzieren
    def finish_func(self):

        """"Die Funktion öffnet ein Abschlussfenster der Inventur. 
        Von diesem Fenster aus können wir den Ordner mit den Inventurergebnissen 
        öffnen und eine Benachrichtigung über den erfolgreichen Abschluss senden."""""

        x = (self.screen_width / 2) - (500 / 2)                                                                         #Fenster platzieren
        y = (self.screen_height / 2) - (650 / 2)
        self.finish = ctk.CTkToplevel(self.artikels_frame)                                                              #Fenser erstellen
        self.finish.geometry(f"{500}x{650}+{int(x)}+{int(y)}")                                                          #Skallieren
        self.finish.grab_set()                                                                                          #Das Fenster wird über den anderen liegen.
        self.finish.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(self.finish, font=ctk.CTkFont("Calibri", weight="bold", size=29), justify="left",                  #Label erstellen
                                  text=f"Inv_{date_today}.csv\n\nwurde erfolgreich erstellt").grid(row=0, padx=25, pady=35)

        def ordner_open(file):

            """"Funktion für Öffnen eines Inventarordners"""""

            subprocess.Popen(f'explorer "{file}"')

        datei_offnen = ctk.CTkButton(self.finish, width=300, height=60, hover_color="#F2C14E", image=self.image_dir, text="Daten öffnen", font=ctk.CTkFont("Calibri", weight="bold", size=29), command=lambda : ordner_open(self.archiv_invent))
        datei_offnen.grid(row=1, pady=(45, 10))                                                                         #"Datei offnen" button
        self.email_button = ctk.CTkButton(self.finish, width=300, height=60, hover_color="#5FAD56", image=self.image_email, text="Per email senden", font=ctk.CTkFont("Calibri", weight="bold", size=29), command=self.sending_message)
        self.email_button.grid(row=2, pady=(10, 10))                                                                    #"Email senden" button
        ctk.CTkButton(self.finish, width=190, height=60, hover_color="#C52233", text="Exit", font=ctk.CTkFont("Calibri", weight="bold", size=25), command=self.finish.destroy).grid(row=3, pady=25)         #Exit button

    def sending_message(self):

        """"Funktion um Emails per Post schicken"""""

        sender = config['email']['sender']                                                                              #Absender email
        sender_pass = config['email']['sender_pass']                                                                    #password
        self.email_button.configure(state="disabled")                                                                   #email button deaktivieren

        # Kontodatei und Server
        creds = Credentials(username=sender, password=sender_pass)                                                      #Verbindung zum Server.
        configur = Configuration(server='mail.argen.de', credentials=creds)

        # Account create
        account = Account(primary_smtp_address=sender, config=configur, autodiscover=False, access_type=DELEGATE)
        directory = config['directory']['archiv_email_dir'] + rf"Inv_{date_today}\upload"
        text = config['email']['emailtext']

        # Brief
        message = Message(account=account,
                    subject=f'Inventur {date_today}',
                    body=f'{text}\n{directory}',
                    to_recipients=self.email_liste)

        message.send()                                                                                                  #Email schicken

        def status_change():
            send_status.configure(image=self.image_access)                                                              #Image von Status Label wechseln

        send_status = ctk.CTkLabel(self.finish, text="", image=self.image_sending)                                      #Status Label erstellen
        send_status.grid(row=4, padx=25, pady=25)                                                                       #Status Label platzieren
        self.after(1000, status_change)                                                                                 #Change status ändern
    def archivs_griding_function(self):

        """"Wir erstellen ein Fenster namens "Archiv", in dem alle vorherigen Inventuren angezeigt werden."""""

        self.main_frame.grid_forget()                                                                                   #Entfernen main frame Fenster
        self.archiv_frame.grid(row=1, sticky="nsew", pady=(10, 0), padx=10)                                             #Archiv Frame platzieren
        self.back_button.configure(command=self.back_to_run)                                                            #Back button Funktion definieren
        files = os.listdir(config['directory']['archiv_dir'])                                                           #Wir erhalten eine Liste der Dateien, die im Ordner "Archiv" gespeichert sind.
        for row, file in enumerate(files):                                                                              #Erstellen button für jede Archivordner
            rows = (row // 4) + 1                                                                                       #Rows griding
            columns = row % 4                                                                                           #Column griding
            ctk.CTkButton(self.archiv_frame, text = file, corner_radius=15, width=250, height=250, fg_color="#50514F",
                                                    image=self.image_folder_mini, compound="top", hover_color="gray50",
                                                    font=ctk.CTkFont("Calibri", size=25, weight="bold"),
            command=lambda file_dir = config['directory']['archiv_dir'] + file + "\\": self.artikel_frame_griding(file_dir, 0)).grid(row=rows, column=columns, padx=20, pady=20)

    def option_function(self):

        """"Es gibt einen Einstellungen-Funktion, der sich immer oben in der Anwendung befindet. 
        Dort kann der Farbmodus und die Textgröße angepasst werden."""""

        x = (self.screen_width / 2) - (1410 / 2) + 100                                                                  #Fenster platzieren
        y = (self.screen_height / 2) - (1200 / 2) + 100
        option_fenster = ctk.CTkToplevel(self)                                                                          #Fenster erstellen
        option_fenster.title("Option")                                                                                  #Fenster Name
        option_fenster.geometry(f"{500}x{500}+{int(x)}+{int(y)}")                                                       #Fenster Größe
        option_fenster.grab_set()                                                                                       #Das Fenster wird über den anderen liegen.

        ctk.CTkLabel(option_fenster,
                     text="Skalierung",
                     font=ctk.CTkFont("Calibri",
                                      size=25,
                                      weight="bold")).grid(row=0, column=0, padx=25, pady=(25, 15), sticky="w")         #"Skalierung" Label erstellen
        scaling_optionmenu = ctk.CTkOptionMenu(option_fenster, values=["80%", "90%", "100%", "110%", "120%"],           #Dropdown Menu erstellen
                                                          command=self.change_scaling_event)
        scaling_optionmenu.set(value="100%")                                                                            #Den Standardwert festlegen.
        scaling_optionmenu.grid(row=0, column=1)                                                                        #Widget platzieren
        switch_var = ctk.StringVar(value="Dark")                                                                        #Erstellung einer Variablen für die Umschaltung des Farbmodus.

        def switch_event():                                                                                             #Funktion für Umschaltung des Farbmodus.
            ctk.set_appearance_mode(switch_var.get())

        ctk.CTkLabel(option_fenster, font=ctk.CTkFont("Calibri", size=25, weight="bold"),                               #"Farbmodus" Label erstellen
                                     text="Farbmodus").grid(row=1, column=0, padx=25, pady=15, sticky="w")
        ctk.CTkSwitch(option_fenster, text="Dark/Light", command=switch_event,                                          #Erstellung einer Schaltfläche zur Umschaltung des Farbmodus.
                                      variable=switch_var, onvalue="Dark", offvalue="Light").grid(row=1, column=1)
    def change_scaling_event(self, new_scaling: str):

        """"Funktion zum Einstellen der Textgröße."""""

        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        ctk.set_widget_scaling(new_scaling_float)

if __name__ == "__main__":
    app = App()
    app.mainloop()