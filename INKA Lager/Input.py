from tkinter import *
import openpyxl
from tkinter import ttk

"""Argen IS Input.exe - die Software, um neue Positionen in Datenbank hinzufügen. 
Die Software basiert auf der Programmiersprache Python. 
Die Informationen werden in eine Excel-Tabelle geschrieben. (Input)"""

win = Tk()

"""die Datei öffnen und lesen, um die Daten auszulesen"""

book = openpyxl.open(r'\\srv-file\Auftrag\IT\ArgenIS\Artikel.xlsx')
sheet = book.active

"""Definieren Globale-varible"""

global label_error
global label_erledig
global label_replay

label_error = Label(text="Bitte fühlen Sie \nalle Felder aus", bg='#2C3639', fg="red", font=('Calibri', 20, 'bold'))
label_erledig = Label(text="Artikel ist \nhinzugefügt", bg='#2C3639', fg="green", font=('Calibri', 20, 'bold'))
label_replay = Label(bg='#2C3639', fg="yellow", font=('Calibri', 20, 'bold'))

"""Hauptfunktion"""

def save():

    replay = []

    for row in range(2, sheet.max_row+1):
        replay.append(sheet[row][0].value)

    max_plus_one = sheet.max_row + 1

    if entry_1.get()[2:16] in replay:
        index = replay.index(entry_1.get()[2:16])
        label_erledig.grid_forget()
        label_error.grid_forget()
        label_replay.config(text=f'Artikel bereit \nexistiert\n {sheet[index+2][2].value}\n{sheet[index+2][3].value}')
        label_replay.grid(row=7, column=2)

    elif entry_1.get()[2:16] not in replay and len(entry_1.get()) > 5 and len(entry_2.get()) >= 5 \
            and combo1.get().isalpha() and combo2.get().isalnum() and position_var.get().isalpha() :

        sheet.cell(max_plus_one, column=1).value = entry_1.get()[2:16]
        sheet.cell(max_plus_one, column=2).value = entry_2.get()
        sheet.cell(max_plus_one, column=3).value = combo1.get() + " - " + combo2.get()
        sheet.cell(max_plus_one, column=4).value = position_var.get()
        book.save(r'\\srv-file\Auftrag\IT\ArgenIS\Artikel.xlsx')
        book.save(r'\\srv-file\Auftrag\IT\Backup_ArgenIS\Artikel_backup.xlsx')
        book.close()
        clear_all()
        label_erledig.grid(row=7, column=2)

    else:
        label_erledig.grid_forget()
        label_replay.grid_forget()
        label_error.grid(row=7, column=2)

"""Funktion um alle Felder löchen (Clear Taste)"""

def clear_all():
    list_ent, label_list, combo_list = [entry_1, entry_2], [label_error, label_replay, label_erledig], [combo1,combo2]
    [i.grid_forget() for i in label_list]
    [i.delete(0, 'end') for i in list_ent]
    [i.set(value='') for i in combo_list]
    position_var.set(0)

"""Erstellen UI(User-Interface)"""

win.title("Argen Input")
win.config(bg='#2C3639')
win.geometry("700x700+600+165")
win.resizable(False, False)

"""Erstellen und platzieren die Elemente auf dem Bildschirm"""

Label(win, bg='#2C3639').grid(row=0,pady=10)
label_1 = Label(text="Artikel", width="20", height="0", anchor="e", bg='#2C3639', fg='white',
                font=('Calibri', 16, 'bold')).grid(row=1, column=0, padx=20)
label_2 = Label(text="Argen Artikelnummer", width="20", height="2", anchor="e", bg='#2C3639', fg='white',
                font=('Calibri', 16, 'bold')).grid(row=3, column=0, padx=20)
label_3 = Label(text="Regal", height="2", anchor="e", bg='#2C3639', fg='white',
                font=('Calibri', 16, 'bold')).grid(row=4, column=0, columnspan=1, stick='we', padx=20)
label_3_1 = Label(text="Platz", height="2", anchor="e", bg='#2C3639', fg='white',
                  font=('Calibri', 16, 'bold')).grid(row=4, column=2, stick='we')
label_4 = Label(text="Position", width="20", height="2", anchor="e", bg='#2C3639',
                fg='white', font=('Calibri', 16, 'bold')).grid(row=5, column=0, columnspan=1, stick='we', padx=20)
label_scan = Label(text='Bitte scannen Sie den Artikel', bg='#2C3639', fg='black',
                   padx=30, font=('Calibri', 9)).grid(row=2, column=1, columnspan=3)


entry_1 = Entry(win, width=10, bg='#3F4E4F', fg='white', font=('Calibri', 16, 'bold'))
entry_2 = Entry(win, width=6, bg='#3F4E4F', fg='white',font=('Calibri', 16, 'bold'))
entry_1.focus_set()

entry_1.grid(row=1, column=1, columnspan=3, stick="we")
entry_2.grid(row=3, column=1)


position_var = StringVar()
position_var.set(0)

radiobutton_1 = Radiobutton(win, text="Vorne", bg='#3F4E4F', fg='black', variable=position_var,
                            value="Vorne", indicatoron=0, font=('Calibri', 20, 'bold'))
radiobutton_2 = Radiobutton(win, text="Hinten", bg='#3F4E4F', fg='black', variable=position_var,
                            value="Hinten", indicatoron=0, font=('Calibri', 20, 'bold'))

radiobutton_1.grid(row=5, column=2)
radiobutton_2.grid(row=5, column=3)


regal = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P',
         'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
place = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]

combo1 = ttk.Combobox(win, values=regal, width=2, font=('Calibri', 20, 'bold'))
combo2 = ttk.Combobox(win, values=place, width=2, font=('Calibri', 20, 'bold'))

combo1.grid(row=4, column=1)
combo2.grid(row=4, column=3)


btn1 = Button(text="Hinzufügen", command=save, width=16, height=3, bg='#A5C9CA', fg='black',
              font=('Calibri', 16, 'bold')).grid(row=6, column=2, pady=20)
btn2 = Button(text="Clear", command=clear_all, width=8, height=2, bg='#3F4E4F', fg='white',
              font=('Calibri', 16)).grid(row=6, column=0, pady=20)


win.grid_columnconfigure(0, minsize=1)
win.grid_columnconfigure(1, minsize=50)
win.grid_columnconfigure(2, minsize=30)
win.grid_columnconfigure(3, minsize=50)
win.grid_columnconfigure(4, minsize=150)

win.grid_rowconfigure(0, minsize=0)
win.grid_rowconfigure(5, minsize=150)
win.grid_rowconfigure(3, minsize=100)


win.mainloop()
