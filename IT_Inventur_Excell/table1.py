import customtkinter as ctk
import configparser
import openpyxl
import custom_treeview as ctv

config = configparser.ConfigParser()
config.read("config.ini", encoding='utf-8')

class Table1(ctk.CTkFrame):

    def __init__(self, master):
        super().__init__(master, fg_color="transparent")
        self.grid_columnconfigure(0, weight=1)

        self.treeview_lager = ctv.CustomTreeView(self, height=35, columns=(
            "column1", "column2", "column3", "column4", "column5"))

        self.tree_scroll_lager = ctk.CTkScrollbar(self, command=self.treeview_lager.yview)
        self.tree_scroll_lager.grid(row=0, column=0, sticky="nse", padx=(0,40), pady=(35, 20))

        self.treeview_lager.configure(yscrollcommand=self.tree_scroll_lager.set)

        self.treeview_lager.heading("#0", text="Item")
        self.treeview_lager.heading("column1", text="Artikel",
                                    command=lambda: self.sort_function("column1", self.treeview_lager, False))
        self.treeview_lager.heading("column2", text="Hersteller",
                                    command=lambda: self.sort_function("column2", self.treeview_lager, False))
        self.treeview_lager.heading("column3", text="Model",
                                    command=lambda: self.sort_function("column3", self.treeview_lager, False))
        self.treeview_lager.heading("column4", text="Seriennummer",
                                    command=lambda: self.sort_function("column4", self.treeview_lager, False))
        self.treeview_lager.heading("column5", text="Bemerkung",
                                    command=lambda: self.sort_function("column5", self.treeview_lager, False))

        self.treeview_lager.column("#0", width=0, minwidth=0, stretch=0)
        self.treeview_lager.column("column1", width=180)
        self.treeview_lager.column("column2", width=180)
        self.treeview_lager.column("column3", width=260)
        self.treeview_lager.column("column4", width=190)
        self.treeview_lager.column("column5", width=224)

        self.treeview_lager.bind("<Double-1>", self.clicker_table_1)

    ### Functions ###
    def sort_function(self, column, table, reverse=False):
        data = [(table.set(child, column), child) for child in table.get_children()]
        data.sort(reverse=reverse)
        for index, (val, child) in enumerate(data):
            table.move(child, '', index)

        table.tag_configure("evenrow", background='gray95')
        table.tag_configure("oddrow", background='white')
        for i, item in enumerate(table.get_children()):
            if i % 2 == 0:
                table.item(item, tags=("evenrow",))
            else:
                table.item(item, tags=("oddrow",))

    def first_table_function(self):
        self.grid(row=1, column=0, sticky="nsew", columnspan=3)

        self.book = openpyxl.open(config['directory']['main_file_dir'])
        self.sheet_1 = self.book.worksheets[0]
        self.max_rows_1 = self.sheet_1.max_row + 1

        self.treeview_lager.delete(*self.treeview_lager.get_children())
        self.treeview_lager.grid_forget()

        self.treeview_lager.tag_configure("oddrow", background="white")
        self.treeview_lager.tag_configure("evenrow", background="gray95")

        for i in range(1, self.max_rows_1):
            for k in range(1, self.sheet_1.max_column + 1):
                if self.sheet_1.cell(row=i, column=k).value != None:
                    pass
                elif self.sheet_1.cell(row=i, column=k).value == None:
                    self.sheet_1.cell(row=i, column=k).value = ""

        for num, record in enumerate(range(1, self.max_rows_1)):
            if self.sheet_1.cell(row=record, column=1).value == "" and self.sheet_1.cell(row=record, column=2).value == "":
                if num % 2 != 0:
                    self.treeview_lager.insert("", "end", iid=num, text="", values=(self.sheet_1.cell(row=record, column=3).value,
                                                                                    self.sheet_1.cell(row=record, column=4).value,
                                                                                    self.sheet_1.cell(row=record, column=5).value,
                                                                                    self.sheet_1.cell(row=record, column=6).value,
                                                                                    self.sheet_1.cell(row=record, column=7).value),
                                                                                                   tags=("oddrow"))

                elif num % 2 == 0:
                    self.treeview_lager.insert("", "end", iid=num, text="", values=(self.sheet_1.cell(row=record, column=3).value,
                                                                                    self.sheet_1.cell(row=record, column=4).value,
                                                                                    self.sheet_1.cell(row=record, column=5).value,
                                                                                    self.sheet_1.cell(row=record, column=6).value,
                                                                                    self.sheet_1.cell(row=record, column=7).value),
                                                                                                   tags=("evenrow"))

        self.treeview_lager.grid(row=0, column=0, sticky="nsew", pady=(35, 20), padx=40)
        self.sort_function("column1", self.treeview_lager, False)

    def clicker_table_1(self, event):

        self.dialog_table1 = ctk.CTkToplevel(self)
        self.dialog_table1.geometry(f"260x290+1200+450")
        self.dialog_table1.resizable(False, False)
        self.dialog_table1.grab_set()
        self.dialog_table1.configure(background="green")
        self.dialog_table1.grid_columnconfigure(0, weight=1)
        self.dialog_table1.grid_columnconfigure(1, weight=1)

        self.artikel_table1_label = ctk.CTkLabel(self.dialog_table1, text="Artikel").grid(row=0, column=0,
                                                                                                    pady=(16, 4),
                                                                                                    sticky="e")
        self.hersteller_table1_label = ctk.CTkLabel(self.dialog_table1, text="Hersteller").grid(row=1,
                                                                                                          column=0,
                                                                                                          pady=4,
                                                                                                          sticky="e")
        self.model_table1_label = ctk.CTkLabel(self.dialog_table1, text="Model").grid(row=2, column=0, pady=4,
                                                                                                sticky="e")
        self.sn_table1_label = ctk.CTkLabel(self.dialog_table1, text="Seriennummer").grid(row=3, column=0,
                                                                                                    pady=4, sticky="e")
        self.bemerkung_table1_label = ctk.CTkLabel(self.dialog_table1, text="Bemerkung").grid(row=4, column=0,
                                                                                                        pady=4,
                                                                                                        sticky="e")
        self.artikel_table1 = ctk.CTkEntry(self.dialog_table1)
        self.artikel_table1.grid(row=0, column=1, pady=(16, 4))
        self.hersteller_table1 = ctk.CTkEntry(self.dialog_table1)
        self.hersteller_table1.grid(row=1, column=1, pady=4)
        self.model_table1 = ctk.CTkEntry(self.dialog_table1)
        self.model_table1.grid(row=2, column=1, pady=4)
        self.sn_table1 = ctk.CTkEntry(self.dialog_table1)
        self.sn_table1.grid(row=3, column=1, pady=4)
        self.bemerkung_table1 = ctk.CTkEntry(self.dialog_table1)
        self.bemerkung_table1.grid(row=4, column=1, pady=4)

        self.selected_table1 = self.treeview_lager.focus()

        self.values_table1 = self.treeview_lager.item(self.selected_table1, 'values')

        self.dialog_table1.title(f"{self.values_table1[0]} {self.values_table1[1]}")

        self.artikel_table1.insert(0, self.values_table1[0])
        self.hersteller_table1.insert(0, self.values_table1[1])
        self.model_table1.insert(0, self.values_table1[2])
        self.sn_table1.insert(0, self.values_table1[3])
        self.bemerkung_table1.insert(0, self.values_table1[4].strip())

        self.confirm_button_table1 = ctk.CTkButton(self.dialog_table1, text="OK",
                                                             command=self.update_record_table_1).grid(row=5, column=1,
                                                                                                      pady=(20, 4))
        self.delete_button_table1 = ctk.CTkButton(self.dialog_table1, text="Löschen", fg_color="#C52233",
                                                            hover_color="#F31B31",
                                                            command=self.delete_command_table1).grid(row=6, column=1,
                                                                                                     pady=4)

    def update_record_table_1(self):
        self.treeview_lager.item(self.selected_table1, text="",
                                 values=(self.artikel_table1.get(), self.hersteller_table1.get(),
                                         self.model_table1.get(),
                                         self.sn_table1.get(),
                                         self.bemerkung_table1.get()))

        for rows in range(1, self.max_rows_1):
            if self.sheet_1.cell(row=rows, column=3).value == self.values_table1[0] \
                    and self.sheet_1.cell(row=rows, column=4).value == self.values_table1[1] \
                    and self.sheet_1.cell(row=rows, column=5).value == self.values_table1[2] \
                    and self.sheet_1.cell(row=rows, column=6).value == self.values_table1[3]:
                self.sheet_1.cell(row=rows, column=3).value = self.artikel_table1.get()
                self.sheet_1.cell(row=rows, column=4).value = self.hersteller_table1.get()
                self.sheet_1.cell(row=rows, column=5).value = self.model_table1.get()
                self.sheet_1.cell(row=rows, column=6).value = self.sn_table1.get()
                self.sheet_1.cell(row=rows, column=7).value = self.bemerkung_table1.get()

        self.book.save(config['directory']['main_file_dir'])
        self.dialog_table1.destroy()

    def delete_command_table1(self):
        self.treeview_lager.item(self.selected_table1, text="",
                                 values=(self.artikel_table1.get(), self.hersteller_table1.get(),
                                         self.model_table1.get(),
                                         self.sn_table1.get(),
                                         self.bemerkung_table1.get()))

        for rows in range(1, self.max_rows_1):
            if self.sheet_1.cell(row=rows, column=3).value == self.values_table1[0] \
                    and self.sheet_1.cell(row=rows, column=4).value == self.values_table1[1] \
                    and self.sheet_1.cell(row=rows, column=5).value == self.values_table1[2] \
                    and self.sheet_1.cell(row=rows, column=6).value == self.values_table1[3]:
                self.sheet_1.delete_rows(rows)
        self.book.save(config['directory']['main_file_dir'])
        self.treeview_lager.delete(self.selected_table1)
        self.dialog_table1.destroy()