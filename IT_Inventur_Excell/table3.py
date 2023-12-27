from tkinter import messagebox
from tkinter.ttk import Style
import customtkinter as ctk
import configparser
import openpyxl
import custom_treeview as ctv

config = configparser.ConfigParser()
config.read("config.ini", encoding='utf-8')

class Table3(ctk.CTkFrame):

    def __init__(self, master):
        super().__init__(master, fg_color="transparent")
        self.grid_columnconfigure(0, weight=1)

        self.style_treeview_style = Style()
        self.style_treeview_style.configure("Treeview", rowheight=25)

        self.treeview_struktur = ctv.CustomTreeView(self, height=35, columns=(
            "column1", "column2", "column3", "column4"))

        self.tree_scroll = ctk.CTkScrollbar(self, command=self.treeview_struktur.yview)
        self.tree_scroll.grid(row=0, column=0, sticky="nse", padx=(0,40), pady=(35, 20))

        self.treeview_struktur.configure(yscrollcommand=self.tree_scroll.set)

        self.treeview_struktur.heading("#0", text="Item")
        self.treeview_struktur.heading("column1", text="Vorname",
                                       command=lambda: self.sort_function("column1", self.treeview_struktur, False))
        self.treeview_struktur.heading("column2", text="Nachname",
                                       command=lambda: self.sort_function("column2", self.treeview_struktur, False))
        self.treeview_struktur.heading("column3", text="Abteilung",
                                       command=lambda: self.sort_function("column3", self.treeview_struktur, False))
        self.treeview_struktur.heading("column4", text="Vorgesetzter",
                                       command=lambda: self.sort_function("column4", self.treeview_struktur, False))

        self.treeview_struktur.column("#0", width=0, minwidth=0, stretch=0)
        self.treeview_struktur.column("column1", width=250)
        self.treeview_struktur.column("column2", width=250)
        self.treeview_struktur.column("column3", width=267)
        self.treeview_struktur.column("column4", width=267)

        self.treeview_struktur.grid(row=2, column=0, padx=(15, 0), columnspan=8)

        self.treeview_struktur.bind("<Double-1>", self.clicker_table_3)

    def sort_function(self, column, table, reverse=False):
        data = [(table.set(child, column), child) for child in table.get_children()]
        data.sort(reverse=reverse)
        for index, (val, child) in enumerate(data):
            table.move(child, '', index)

        table.tag_configure("evenrow", background='gray95')
        table.tag_configure("oddrow", background='white')
        for i, item in enumerate(table.get_children()):
            if i % 2 == 0:
                table.item(item, tags=("evenrow",))
            else:
                table.item(item, tags=("oddrow",))

    def third_table_funktion(self):

        self.grid(row=1, column=0, sticky="nsew", columnspan=3)
        self.book = openpyxl.open(config['directory']['main_file_dir'])
        self.sheet_2 = self.book.worksheets[1]
        self.max_rows_2 = self.sheet_2.max_row + 1

        self.treeview_struktur.delete(*self.treeview_struktur.get_children())
        self.treeview_struktur.grid_forget()

        self.treeview_struktur.tag_configure("oddrow", background="white")
        self.treeview_struktur.tag_configure("evenrow", background="gray95")

        for num, record in enumerate(range(1, self.max_rows_2)):
            if num % 2 != 0:
                self.treeview_struktur.insert("", "end", iid=num, text="",
                                           values=(self.sheet_2.cell(row=record, column=1).value,
                                                   self.sheet_2.cell(row=record, column=2).value,
                                                   self.sheet_2.cell(row=record, column=3).value,
                                                   self.sheet_2.cell(row=record, column=4).value),
                                                   tags=("oddrow"))
            elif num % 2 == 0:
                self.treeview_struktur.insert("", "end", iid=num, text="",
                                           values=(self.sheet_2.cell(row=record, column=1).value,
                                                   self.sheet_2.cell(row=record, column=2).value,
                                                   self.sheet_2.cell(row=record, column=3).value,
                                                   self.sheet_2.cell(row=record, column=4).value),
                                                   tags=("evenrow"))


        self.treeview_struktur.grid(row=0, column=0, sticky="nsew", pady=(35, 20), padx=40)
        self.sort_function("column1", self.treeview_struktur, False)

    def clicker_table_3(self, event):
        self.dialog_table3 = ctk.CTkToplevel(self)
        self.dialog_table3.geometry("260x290+1200+450")
        self.dialog_table3.resizable(False, False)
        self.dialog_table3.grab_set()
        self.dialog_table3.grid_columnconfigure(0, weight=1)
        self.dialog_table3.grid_columnconfigure(1, weight=1)

        self.vorname_table_3 = ctk.CTkLabel(self.dialog_table3, text="Vorname").grid(row=0, column=0,
                                                                                               pady=(16, 4),
                                                                                               sticky="e")
        self.nachname_table_3 = ctk.CTkLabel(self.dialog_table3, text="Nachname").grid(row=1,
                                                                                                 column=0,
                                                                                                 pady=4,
                                                                                                 sticky="e")
        self.abteilung_table_3 = ctk.CTkLabel(self.dialog_table3, text="Abteilung").grid(row=2, column=0,
                                                                                                   pady=4,
                                                                                                   sticky="e")
        self.vorgesetzter_table_3 = ctk.CTkLabel(self.dialog_table3, text="Vorgesetzter").grid(row=3,
                                                                                                         column=0,
                                                                                                         pady=4,
                                                                                                         sticky="e")

        self.vorname_table3 = ctk.CTkEntry(self.dialog_table3)
        self.vorname_table3.grid(row=0, column=1, pady=(16, 4))
        self.nachname_table3 = ctk.CTkEntry(self.dialog_table3)
        self.nachname_table3.grid(row=1, column=1, pady=4)
        self.abteilung_table3 = ctk.CTkEntry(self.dialog_table3)
        self.abteilung_table3.grid(row=2, column=1, pady=4)
        self.vorgesetzter_table3 = ctk.CTkEntry(self.dialog_table3)
        self.vorgesetzter_table3.grid(row=3, column=1, pady=4)

        self.selected_table3 = self.treeview_struktur.focus()
        self.values_table3 = self.treeview_struktur.item(self.selected_table3, 'values')

        self.dialog_table3.title(f"{self.values_table3[0]} {self.values_table3[1]}")

        self.vorname_table3.insert(0, self.values_table3[0])
        self.nachname_table3.insert(0, self.values_table3[1])
        self.abteilung_table3.insert(0, self.values_table3[2])
        self.vorgesetzter_table3.insert(0, self.values_table3[3])

        self.confirm_button_table3 = ctk.CTkButton(self.dialog_table3, text="OK",
                                                             command=self.update_record_table_3).grid(row=5, column=1,
                                                                                                      pady=(30, 4))

        self.delete_button_table3 = ctk.CTkButton(self.dialog_table3, text="Löschen", fg_color="#C52233",
                                                            hover_color="#F31B31",
                                                            command=self.delete_command_table3).grid(row=6, column=1,
                                                                                                     pady=4)

    def update_record_table_3(self):
        self.treeview_struktur.item(self.selected_table3, text="",
                                    values=(self.vorname_table3.get(),
                                            self.nachname_table3.get(),
                                            self.abteilung_table3.get(),
                                            self.vorgesetzter_table3.get()))

        for rows in range(1, self.max_rows_2):
            if self.sheet_2.cell(row=rows, column=1).value == self.values_table3[0] \
                    and self.sheet_2.cell(row=rows, column=2).value == self.values_table3[1] \
                    and self.sheet_2.cell(row=rows, column=3).value == self.values_table3[2] \
                    and self.sheet_2.cell(row=rows, column=4).value == self.values_table3[3]:
                self.sheet_2.cell(row=rows, column=1).value = self.vorname_table3.get()
                self.sheet_2.cell(row=rows, column=2).value = self.nachname_table3.get()
                self.sheet_2.cell(row=rows, column=3).value = self.abteilung_table3.get()
                self.sheet_2.cell(row=rows, column=4).value = self.vorgesetzter_table3.get()

        self.book.save(config['directory']['main_file_dir'])
        self.dialog_table3.destroy()

    def delete_command_table3(self):

        delete_bestaetigen = messagebox.askyesno("Bitte bestätigen",
                                                 f"Sind Sie sicher, dass Sie den Mitarbeiter {self.values_table3[0]} {self.values_table3[1]} löschen möchten?")

        if delete_bestaetigen:
            self.treeview_struktur.item(self.selected_table3, text="",
                                     values=(self.vorname_table3.get(), self.nachname_table3.get(),
                                             self.abteilung_table3.get(), self.vorgesetzter_table3.get()))

            for rows in range(1, self.max_rows_2):
                if self.sheet_2.cell(row=rows, column=1).value == self.values_table3[0] \
                        and self.sheet_2.cell(row=rows, column=2).value == self.values_table3[1] \
                        and self.sheet_2.cell(row=rows, column=3).value == self.values_table3[2] \
                        and self.sheet_2.cell(row=rows, column=4).value == self.values_table3[3]:
                    self.sheet_2.delete_rows(rows)

            self.book.save(config['directory']['main_file_dir'])
            self.treeview_struktur.delete(self.selected_table3)
            self.dialog_table3.destroy()

        else:
            pass