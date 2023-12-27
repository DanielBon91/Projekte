import customtkinter as ctk
from datetime import date
from PIL import Image
import openpyxl
import configparser

config = configparser.ConfigParser()
config.read("config.ini", encoding='utf-8')
img_dir = config['directory']['images_dir']
date_today = date.today().strftime("%d.%m.%Y")

class FirstFrame(ctk.CTkFrame):

    def __init__(self, master):
        super().__init__(master, fg_color="transparent")

        self.grid_columnconfigure(0, minsize=150)
        self.grid_columnconfigure(1, weight=0)
        self.grid_columnconfigure(2, weight=1)
        self.grid_columnconfigure(3, minsize=145)
        self.grid_rowconfigure(0, minsize=145)

        # Images create
        self.image_access = ctk.CTkImage(Image.open(img_dir + "access.png"), size=(70, 70))
        self.image_error_artikel = ctk.CTkImage(Image.open(img_dir + "error.png"), size=(30, 30))
        self.image_error_confirm = ctk.CTkImage(Image.open(img_dir + "error.png"), size=(50, 50))

        # Labels create
        self.artikel_label = ctk.CTkLabel(self, text="Artikel", font=ctk.CTkFont("Calibri", size=22, weight="bold"))
        self.artikel_label.grid(row=1, column=1, pady=15, sticky="w")
        self.hersteller_label = ctk.CTkLabel(self, text="Hersteller", font=ctk.CTkFont("Calibri", size=22, weight="bold"))
        self.hersteller_label.grid(row=2, column=1, pady=15, sticky="w")
        self.model_label = ctk.CTkLabel(self, text="Model", font=ctk.CTkFont("Calibri", size=22, weight="bold"))
        self.model_label.grid(row=3, column=1, pady=15, sticky="w")
        self.sn_label = ctk.CTkLabel(self, text="Seriennummer", font=ctk.CTkFont("Calibri", size=22, weight="bold"))
        self.sn_label.grid(row=4, column=1, padx=(0, 40), pady=15, sticky="w")
        self.datum_label = ctk.CTkLabel(self, text="Datum", font=ctk.CTkFont("Calibri", size=22, weight="bold"))
        self.datum_label.grid(row=5, column=1, pady=15, sticky="w")
        self.bemerkung_label = ctk.CTkLabel(self, text="Bemerkung", font=ctk.CTkFont("Calibri", size=22, weight="bold"))
        self.bemerkung_label.grid(row=6, column=1, pady=15, sticky="w")
        self.label_access = ctk.CTkLabel(self, text="", image=self.image_access, font=ctk.CTkFont("Calibri", size=44))
        self.label_error_artikel = ctk.CTkLabel(self, text="", image=self.image_error_artikel)
        self.label_error_confirm = ctk.CTkLabel(self, text="", image=self.image_error_confirm)

        # Entry boxes create
        self.artikel_entry = ctk.CTkEntry(self, corner_radius=0, font=ctk.CTkFont("Calibri", size=22))
        self.artikel_entry.grid(row=1, column=2, pady=15, sticky="we")
        self.artikel_entry.bind("<FocusOut>", self.change_artikel)
        self.hersteller_entry = ctk.CTkEntry(self, corner_radius=0, font=ctk.CTkFont("Calibri", size=22))
        self.hersteller_entry.grid(row=2, column=2, pady=15, sticky="we")
        self.model_entry = ctk.CTkEntry(self, corner_radius=0, font=ctk.CTkFont("Calibri", size=22))
        self.model_entry.grid(row=3, column=2, pady=15, sticky="we")
        self.sn_entry = ctk.CTkEntry(self, corner_radius=0, font=ctk.CTkFont("Calibri", size=22))
        self.sn_entry.grid(row=4, column=2, pady=15, sticky="we")
        self.datum_entry = ctk.CTkEntry(self, corner_radius=0, font=ctk.CTkFont("Calibri", size=22))
        self.datum_entry.insert("0", date_today)
        self.datum_entry.grid(row=5, column=2, pady=15, sticky="we")
        self.bemerkung_entry = ctk.CTkTextbox(self, corner_radius=0, height=100, font=ctk.CTkFont("Calibri", size=22))
        self.bemerkung_entry.grid(row=6, column=2, pady=15, sticky="we")

        # Buttons create
        self.button_confirm = ctk.CTkButton(self, text="Confirm", corner_radius=0, height=75, width=265,
                                            font=ctk.CTkFont("Calibri", size=34), hover_color="#5fad56",
                                            command=self.first_frame_writing_data)
        self.button_confirm.grid(row=7, column=2, pady=(55, 15))
        self.button_clear = ctk.CTkButton(self, text="Clear all", corner_radius=0, height=45, width=210,
                                            fg_color="gray", hover_color="#C52233",
                                            font=ctk.CTkFont("Calibri", size=14), command=self.first_frame_clear_all)
        self.button_clear.grid(row=8, column=2, pady=(55, 15))

    ### Functions ###

    def change_artikel(self, event):

        #todo: erstellen List im List mit der Werten
        werte_list = []
        for werte in config['wert']['werte'].split(':'):
            sub_list_wert = []
            sub_list_wert.append(werte)
            werte_list.append(sub_list_wert)                  #[['Smartphone'], ['Bildschirm'], ['Laptop'], ['Transponderchip']...]

        #todo: falsche Werten zum bestimmte Liste hinzu
        for wert_num in range(len(werte_list)):
            for falsches_wert in config['wert'][werte_list[wert_num][0]].split(':'):
                werte_list[wert_num].append(falsches_wert)    #[['Smartphone',...], ['Bildschirm',...], ['Laptop',...], ['Transponderchip',...],...]

        #todo: die Werte werden durch die richtigen ersetzt
        for num, value_list in enumerate(werte_list):
            for value in value_list:
                if self.artikel_entry.get().lower() in value:
                    self.artikel_entry.delete(0, "end")
                    self.artikel_entry.insert(0, werte_list[num][0])

    def first_frame_writing_data(self):

        book = openpyxl.open(config['directory']['main_file_dir'])
        sheet = book.worksheets[0]
        max_rows = sheet.max_row + 1

        if len(self.artikel_entry.get()) > 0:
            sheet.cell(max_rows, column=1).value = None
            sheet.cell(max_rows, column=2).value = None
            sheet.cell(max_rows, column=3).value = self.artikel_entry.get().capitalize().strip()
            sheet.cell(max_rows, column=4).value = self.hersteller_entry.get().capitalize().capitalize().strip()
            sheet.cell(max_rows, column=5).value = self.model_entry.get().strip()
            sheet.cell(max_rows, column=6).value = str(self.sn_entry.get().strip())
            sheet.cell(max_rows, column=7).value = self.bemerkung_entry.get('0.0', 'end').strip()
            sheet.cell(max_rows, column=8).value = self.datum_entry.get().strip()

            self.artikel_entry.delete(0, "end")
            self.hersteller_entry.delete(0, "end")
            self.model_entry.delete(0, "end")
            self.sn_entry.delete(0, "end")
            self.datum_entry.delete(0, "end")
            self.datum_entry.insert(0, date_today)
            self.bemerkung_entry.delete("0.0", "end")
            self.label_access.grid(row=7, column=2, padx=(450,0), pady=(35, 0))
        else:
            self.label_access.grid_forget()
            self.label_error_artikel.grid(row=1, column=3)
            self.label_error_confirm.grid(row=7, column=2, padx=(450,0), pady=(35, 0))

        book.save(config['directory']['main_file_dir'])

        def label_delete(event):
            self.label_error_artikel.grid_forget()
            self.label_error_confirm.grid_forget()
            self.label_access.grid_forget()
            self.datum_entry.delete(0, "end")
            self.datum_entry.insert(0, date_today)
        self.artikel_entry.bind("<KeyRelease>", label_delete)

    def first_frame_clear_all(self):
        self.artikel_entry.delete(0, "end")
        self.hersteller_entry.delete(0, "end")
        self.model_entry.delete(0, "end")
        self.sn_entry.delete(0, "end")
        self.datum_entry.delete(0, "end")
        self.datum_entry.insert(0, date_today)
        self.bemerkung_entry.delete("0.0", "end")
        self.label_error_artikel.grid_forget()
        self.label_error_confirm.grid_forget()
        self.label_access.grid_forget()
        self.artikel_entry.focus_set()