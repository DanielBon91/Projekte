import os
import subprocess
import customtkinter as ctk
import custom_treeview as ctv
from PIL import Image
import openpyxl
import configparser
from docxtpl import DocxTemplate

config = configparser.ConfigParser()
config.read("config.ini", encoding='utf-8')
img_dir = config['directory']['images_dir']

class SecondFrame(ctk.CTkFrame):

    def __init__(self, master):
        super().__init__(master, fg_color="transparent")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # Images create
        self.image_uebergabe = ctk.CTkImage(Image.open(img_dir + "uebergabe.png"), size=(35, 35))
        self.image_plus = ctk.CTkImage(Image.open(img_dir + "plus.png"), size=(35, 35))
        self.image_minus = ctk.CTkImage(Image.open(img_dir + "minus.png"), size=(35, 35))
        self.image_pfeil = ctk.CTkImage(Image.open(img_dir + "pfeil.png"), size=(35, 35))
        self.word = ctk.CTkImage(Image.open(fr"{img_dir}//word.png"), size=(35, 35))

        # Create upper table
        self.lager_table = ctv.CustomTreeView(self, columns=("column1", "column2", "column3", "column4", "column5"))
        self.scrollbar_lager_table = ctk.CTkScrollbar(self, command=self.lager_table.yview)
        self.scrollbar_lager_table.grid(row=0, column=0, sticky="nse", padx=(0,40), pady=(35, 20))
        self.lager_table.configure(yscrollcommand=self.scrollbar_lager_table.set)
        self.lager_table.heading("#0", text="")
        self.lager_table.heading("column1", text="Artikel", command=lambda: self.sort_function("column1", self.lager_table, False))
        self.lager_table.heading("column2", text="Hersteller", command=lambda: self.sort_function("column2", self.lager_table, False))
        self.lager_table.heading("column3", text="Model", command=lambda: self.sort_function("column3", self.lager_table, False))
        self.lager_table.heading("column4", text="Seriennummer", command=lambda: self.sort_function("column4", self.lager_table, False))
        self.lager_table.heading("column5", text="Bemerkung", command=lambda: self.sort_function("column5", self.lager_table, False))

        self.lager_table.column("#0", width=0, minwidth=0, stretch=0)
        self.lager_table.column("column1", width=120)
        self.lager_table.column("column2", width=120)
        self.lager_table.column("column3")
        self.lager_table.column("column4")
        self.lager_table.column("column5")
        self.lager_table.grid(row=0, column=0, sticky="nsew", pady=(35, 20), padx=40)

        # Create bottom table
        self.empty_table = ctv.CustomTreeView(self, height=10, columns=( "column1", "column2", "column3", "column4", "column5"))
        self.empty_table.heading("#0", text="")
        self.empty_table.heading("column1", text="Artikel", command=lambda: self.sort_function("column1", self.empty_table, False))
        self.empty_table.heading("column2", text="Hersteller", command=lambda: self.sort_function("column2", self.empty_table, False))
        self.empty_table.heading("column3", text="Model", command=lambda: self.sort_function("column3", self.empty_table,  False))
        self.empty_table.heading("column4", text="Seriennummer", command=lambda: self.sort_function("column4", self.empty_table, False))
        self.empty_table.heading("column5", text="Bemerkung", command=lambda: self.sort_function("column5", self.empty_table, False))

        self.empty_table.column("#0", width=0, minwidth=0, stretch=0)
        self.empty_table.column("column1", width=120)
        self.empty_table.column("column2", width=120)
        self.empty_table.column("column3")
        self.empty_table.column("column4")
        self.empty_table.column("column5")
        self.empty_table.grid(row=2, column=0, sticky="nsew", padx=40, pady=20)

        # Button create
        self.zuweisen_button = ctk.CTkButton(self, text="Die Waren übergebe an...", width=350, height=80,
                                             state="disabled", image=self.image_uebergabe, command=self.uebergabe_function,
                                             font=ctk.CTkFont(size=25))
        self.zuweisen_button.grid(row=4, column=0, pady=(0, 20))
        self.plus_button = ctk.CTkButton(self, text="", fg_color="#C52233", corner_radius=0, hover_color="#F31B31",
                                         command=self.plus_function, image=self.image_plus, height=50)
        self.plus_button.grid(row=1, column=0, sticky="w", padx=(100, 0))
        self.minus_button = ctk.CTkButton(self, text="", fg_color="#399E5A", corner_radius=0, hover_color="#328E3D",
                                             command=self.minus_function, image=self.image_minus, height=50)
        self.minus_button.grid(row=1, column=0, sticky="e", padx=(0, 100))

        self.lager_table.bind("<Double-1>", self.plus_click)
        self.empty_table.bind("<Double-1>", self.minus_click)


    ### Functions ###

    def second_frame_lager_tabelle(self):
        self.grid(row=0, column=1, sticky="nsew")
        self.book = openpyxl.open(config['directory']['main_file_dir'])
        self.sheet_1 = self.book.worksheets[0]
        self.max_rows_1 = self.sheet_1.max_row + 1

        self.lager_table.delete(*self.lager_table.get_children())
        self.lager_table.grid_forget()

        self.lager_table.tag_configure("oddrow", background="white")
        self.lager_table.tag_configure("evenrow", background="gray95")

        for i in range(1, self.max_rows_1):
            for k in range(1, self.sheet_1.max_column + 1):
                if self.sheet_1.cell(row=i, column=k).value != None:
                    pass
                elif self.sheet_1.cell(row=i, column=k).value == None:
                    self.sheet_1.cell(row=i, column=k).value = ""

        for num, record in enumerate(range(1, self.max_rows_1)):
            if self.sheet_1.cell(row=record, column=1).value == "" and self.sheet_1.cell(row=record, column=2).value == "":
                if num % 2 != 0:
                    self.lager_table.insert("", "end", iid=num, text="", values=(self.sheet_1.cell(row=record, column=3).value,
                                                                                 self.sheet_1.cell(row=record, column=4).value,
                                                                                 self.sheet_1.cell(row=record, column=5).value,
                                                                                 self.sheet_1.cell(row=record, column=6).value,
                                                                                 self.sheet_1.cell(row=record, column=7).value),
                                                                                                   tags=("oddrow"))

                elif num % 2 == 0:
                    self.lager_table.insert("", "end", iid=num, text="", values=(self.sheet_1.cell(row=record, column=3).value,
                                                                                 self.sheet_1.cell(row=record, column=4).value,
                                                                                 self.sheet_1.cell(row=record, column=5).value,
                                                                                 self.sheet_1.cell(row=record, column=6).value,
                                                                                 self.sheet_1.cell(row=record, column=7).value),
                                                                                                   tags=("evenrow"))

        self.lager_table.grid(row=0, column=0, sticky="nsew", pady=(35, 20), padx=40)
        self.sort_function("column1", self.lager_table, False)

    def sort_function(self, column, table, reverse=False):
        data = [(table.set(child, column), child) for child in table.get_children()]
        data.sort(reverse=reverse)
        for index, (val, child) in enumerate(data):
            table.move(child, '', index)

        table.tag_configure("evenrow", background='gray95')
        table.tag_configure("oddrow", background='white')
        for i, item in enumerate(table.get_children()):
            if i % 2 == 0:
                table.item(item, tags=("evenrow",))
            else:
                table.item(item, tags=("oddrow",))

    def plus_click(self, event):
        self.plus_function()

    def plus_function(self):
        for rows in self.lager_table.selection():
            self.empty_table.insert("", "end", text="", values=(self.lager_table.item(rows, 'values')[0],
                                                                self.lager_table.item(rows, 'values')[1],
                                                                self.lager_table.item(rows, 'values')[2],
                                                                self.lager_table.item(rows, 'values')[3],
                                                                self.lager_table.item(rows, 'values')[4]))
            self.lager_table.delete(self.lager_table.selection()[0])

        if len(self.empty_table.get_children()) > 0:
            self.zuweisen_button.configure(state="normal")

    def minus_click(self, event):
        self.minus_function()

    def minus_function(self):
        for rows in self.empty_table.selection():
            self.lager_table.insert("", "end", text="", values=(self.empty_table.item(rows, 'values')[0],
                                                                self.empty_table.item(rows, 'values')[1],
                                                                self.empty_table.item(rows, 'values')[2],
                                                                self.empty_table.item(rows, 'values')[3],
                                                                self.empty_table.item(rows, 'values')[4]))
            self.empty_table.delete(self.empty_table.selection()[0])

        if len(self.empty_table.get_children()) == 0:
            self.zuweisen_button.configure(state="disabled")

    def uebergabe_function(self):
        self.dialog_mitarbeiter = ctk.CTkToplevel(self)
        self.dialog_mitarbeiter.title("Bitte auswählen")
        self.dialog_mitarbeiter.geometry("800x750+1030+250")
        self.dialog_mitarbeiter.resizable(False, False)
        self.dialog_mitarbeiter.grab_set()

        self.dialog_mitarbeiter.grid_columnconfigure(0, weight=1)
        self.dialog_mitarbeiter.grid_columnconfigure(1, weight=1)

        self.dialog_mitarbeiter_label_vorname = ctk.CTkLabel(self.dialog_mitarbeiter, text="Vorname", font=ctk.CTkFont("Calibri", weight="bold", size=25))
        self.dialog_mitarbeiter_label_vorname.grid(row=0, column=0, sticky="w", padx=(35, 0), pady=15)
        self.dialog_mitarbeiter_label_nachname = ctk.CTkLabel(self.dialog_mitarbeiter, text="Nachname", font=ctk.CTkFont("Calibri", weight="bold", size=25))
        self.dialog_mitarbeiter_label_nachname.grid(row=0, column=1, sticky="w", padx=(35, 0), pady=15)

        self.sheet_2 = self.book.worksheets[1]
        self.max_rows_2 = self.sheet_2.max_row + 1

        vorname_list = []
        for i in range(1, self.max_rows_2):
            vorname_list.append(self.sheet_2.cell(row=i, column=1).value)
        sort_vorname_list = sorted(set(vorname_list))

        self.dialog_mitarbeiter_box_vorname = ctk.CTkOptionMenu(self.dialog_mitarbeiter, width=200, values=sort_vorname_list,
                                                                          command=self.vorwahl_nachname,
                                                                          font=ctk.CTkFont("Calibri", size=21, weight="bold"),
                                                                          dropdown_font=ctk.CTkFont("Calibri", size=19, weight="bold"))
        self.dialog_mitarbeiter_box_vorname.grid(row=1, column=0)
        self.dialog_mitarbeiter_box_vorname.set("Bitte auswählen")
        self.dialog_mitarbeiter_box_nachname = ctk.CTkOptionMenu(self.dialog_mitarbeiter, state="disabled",
                                                                          width=200, font=ctk.CTkFont("Calibri", size=21, weight="bold"),
                                                                          dropdown_font=ctk.CTkFont("Calibri", size=19, weight="bold"))
        self.dialog_mitarbeiter_box_nachname.grid(row=1, column=1)
        self.dialog_mitarbeiter_box_nachname.set("Bitte auswählen")

    def vorwahl_nachname(self, vorname):
        self.dialog_mitarbeiter_box_nachname.configure(state="normal")
        self.dialog_mitarbeiter_box_nachname.set("Bitte auswählen")

        nachname_list = []
        for i in range(1, self.max_rows_2):
            if self.sheet_2.cell(row=i, column=1).value == vorname:
                nachname_list.append(self.sheet_2.cell(row=i, column=2).value)
        nachname = self.dialog_mitarbeiter_box_nachname.get()
        self.dialog_mitarbeiter_box_nachname.configure(values=nachname_list, command=lambda nachname = nachname: self.confirm_function(vorname, nachname))

    def confirm_function(self, vorname, nachname):
        ctk.CTkLabel(self.dialog_mitarbeiter, font=ctk.CTkFont(size=25, weight="bold"),
                                              text=(f"{vorname} {nachname} bekomt die Waren:")).grid(row=2, column=0,
                                              columnspan=2, padx=(50, 0), pady=(45, 15), sticky="w")

        for row in self.empty_table.get_children():
            sofort_label = ' '.join(str(x) for x in self.empty_table.item(row)['values'][:4])
            ctk.CTkLabel(self.dialog_mitarbeiter, font=ctk.CTkFont(size=18), text=f"""- {sofort_label}""").grid(column=0, columnspan=2, sticky="w", padx=(75, 0))

        ctk.CTkLabel(self.dialog_mitarbeiter, text="Übergabedatum:", font=ctk.CTkFont(size=19, weight="bold")).grid(row=14, column=0, columnspan=3,
                                                                                           sticky="w", padx=(60, 0), pady=20)
        self.data_entry = ctk.CTkEntry(self.dialog_mitarbeiter, height=35, placeholder_text="dd.mm.YYYY", font=ctk.CTkFont(size=19))
        self.data_entry.grid(row=14, column=0, sticky="w", padx=(230, 0), pady=20, columnspan=3)
        self.data_get = ctk.CTkButton(self.dialog_mitarbeiter, width=45, text="", image=self.image_pfeil, command=lambda: self.button_active(self.data_entry.get()))
        self.data_get.grid(row=14, column=0, columnspan=3)

        for i in range(1, self.max_rows_2):
            if self.sheet_2.cell(row=i, column=1).value == vorname and self.sheet_2.cell(row=i, column=2).value == nachname:
                abteilung = self.sheet_2.cell(row=i, column=3).value
                chef = self.sheet_2.cell(row=i, column=4).value

        self.top_level_confirm_button = ctk.CTkButton(self.dialog_mitarbeiter, width=200, height=45,
                                                      text="Bestätigen", state="disabled",
                                                      font=ctk.CTkFont(size=21, weight="bold"),
                                                      command=lambda: self.bestaetigung_command(vorname, nachname, abteilung, chef))
        self.top_level_confirm_button.grid(row=15, column=0, columnspan=3, pady=45, sticky="s")

    def button_active(self, date):
        if len(date) == 10:
            self.top_level_confirm_button.configure(state="normal")
        else:
            pass

    def bestaetigung_command(self, vorname, nachname, abteilung, chef):
        dictionary_uebergabe = {'name': vorname,
                                'nachname': nachname,
                                'abteilung': abteilung,
                                'chef': chef}
        for value in self.empty_table.get_children():

            artikel = self.empty_table.item(value)['values'][0]
            hersteller = self.empty_table.item(value)['values'][1]
            model = self.empty_table.item(value)['values'][2]
            seriennummer = self.empty_table.item(value)['values'][3]

            for rows in range(1, self.max_rows_1):
                if self.sheet_1.cell(row=rows, column=3).value == artikel \
                        and self.sheet_1.cell(row=rows, column=4).value == hersteller \
                        and self.sheet_1.cell(row=rows, column=5).value == str(model) \
                        and self.sheet_1.cell(row=rows, column=6).value == str(seriennummer):
                    self.sheet_1.cell(row=rows, column=1).value = vorname
                    self.sheet_1.cell(row=rows, column=2).value = nachname

                    self.book.save(config['directory']['main_file_dir'])

        for num, rows in enumerate(self.empty_table.get_children()):
            art_name = ' '.join(str(value) for value in self.empty_table.item(rows)['values'][:3])
            serial_num_word = self.empty_table.item(rows)['values'][3]
            dictionary_uebergabe[f'art{num}'] = art_name
            dictionary_uebergabe[f'sn{num}'] = serial_num_word
            dictionary_uebergabe[f'dat{num}'] = self.data_entry.get()

        self.empty_table.delete(*self.empty_table.get_children())

        default_word_datei = (config['directory']['uebergabeprotokoll_default'])
        word_datei = DocxTemplate(default_word_datei)
        word_datei.render(dictionary_uebergabe)

        count_name = 0
        user_uebergabe_word_datei = (config['directory']['uebergabeprotokoll_directory'])
        files = os.listdir(user_uebergabe_word_datei)

        vollname = f'{vorname}_{nachname}'

        for name in files:
            if vollname in name:
                count_name += 1

        if count_name == 0:
            file_name = f"Übergabeprotokoll_{vorname}_{nachname}.docx"
            end_word_directory = user_uebergabe_word_datei + file_name
            word_datei.save(end_word_directory)
        else:
            file_name = f"Übergabeprotokoll_{vorname}_{nachname}_({count_name + 1}).docx"
            end_word_directory = user_uebergabe_word_datei + file_name
            word_datei.save(end_word_directory)

        self.bestaetigung_button = ctk.CTkButton(self.dialog_mitarbeiter, text="Word", image=self.word,
                                                           command=lambda: self.open(end_word_directory))
        self.bestaetigung_button.grid(row=16, column=0, columnspan=2)

        self.book.save(config['directory']['main_file_dir'])

    def open(self, end_word):
        subprocess.Popen(['start', end_word], shell=True)
        self.dialog_mitarbeiter.destroy()
        self.zuweisen_button.configure(state="disabled")