from tkinter import messagebox
import customtkinter as ctk
import configparser
import openpyxl
import custom_treeview as ctv
from PIL import Image

config = configparser.ConfigParser()
config.read("config.ini", encoding='utf-8')

class Table2(ctk.CTkFrame):

    def __init__(self, master):
        super().__init__(master, fg_color="transparent")
        self.grid_columnconfigure(0, weight=1)

        img_dir = config['directory']['images_dir']
        self.image_rueckgabe = ctk.CTkImage(Image.open(fr"{img_dir}//rueckgabe.png"), size=(35, 35))

        self.treeview_inventar = ctv.CustomTreeView(self, height=30, columns=(
            "column1", "column2", "column3", "column4", "column5", "column6", "column7"))

        self.tree_scroll_invent = ctk.CTkScrollbar(self, command=self.treeview_inventar.yview)
        self.tree_scroll_invent.grid(row=0, column=0, sticky="nse", padx=(0,40), pady=(35, 20))

        self.treeview_inventar.configure(yscrollcommand=self.tree_scroll_invent.set)

        self.treeview_inventar.heading("#0", text="Item")
        self.treeview_inventar.heading("column1", text="Vorname",
                                       command=lambda: self.sort_function("column1", self.treeview_inventar, False))
        self.treeview_inventar.heading("column2", text="Nachname",
                                       command=lambda: self.sort_function("column2", self.treeview_inventar, False))
        self.treeview_inventar.heading("column3", text="Artikel",
                                       command=lambda: self.sort_function("column3", self.treeview_inventar, False))
        self.treeview_inventar.heading("column4", text="Hersteller",
                                       command=lambda: self.sort_function("column4", self.treeview_inventar, False))
        self.treeview_inventar.heading("column5", text="Model",
                                       command=lambda: self.sort_function("column5", self.treeview_inventar, False))
        self.treeview_inventar.heading("column6", text="Seriennummer",
                                       command=lambda: self.sort_function("column6", self.treeview_inventar, False))
        self.treeview_inventar.heading("column7", text="Bemerkung",
                                       command=lambda: self.sort_function("column7", self.treeview_inventar, False))

        self.treeview_inventar.column("#0", width=0, minwidth=0, stretch=0)
        self.treeview_inventar.column("column1", width=130)
        self.treeview_inventar.column("column2", width=130)
        self.treeview_inventar.column("column3", width=180)
        self.treeview_inventar.column("column4", width=120)
        self.treeview_inventar.column("column5", width=190)
        self.treeview_inventar.column("column6", width=169)
        self.treeview_inventar.column("column7", width=190)

        self.treeview_inventar.bind("<Double-1>", self.clicker_table_2)

        self.rueckgabe_button = ctk.CTkButton(self, width=300, height=70, corner_radius=0,
                                              text="Rückgabe machen", image=self.image_rueckgabe,
                                              fg_color="#328E3D", hover_color="#399E5A",
                                              font=ctk.CTkFont(size=21, weight="bold"),
                                              command=self.rueckgabe_fuction).grid(row=1, column=0, padx=40, pady=25, sticky="w")

        self.search = ctk.CTkEntry(self, corner_radius=0, width=200)
        self.search.grid(row=1, column=0, padx=15, pady=25)
        self.search.bind("<KeyRelease>", self.search_funktion_event)

    ### Functions ###
    def sort_function(self, column, table, reverse=False):
        data = [(table.set(child, column), child) for child in table.get_children()]
        data.sort(reverse=reverse)
        for index, (val, child) in enumerate(data):
            table.move(child, '', index)

        table.tag_configure("evenrow", background='gray95')
        table.tag_configure("oddrow", background='white')
        for i, item in enumerate(table.get_children()):
            if i % 2 == 0:
                table.item(item, tags=("evenrow",))
            else:
                table.item(item, tags=("oddrow",))

    def second_table_function(self):
        self.grid(row=1, column=0, sticky="nsew", columnspan=3)
        self.search.delete(0, "end")
        self.book = openpyxl.open(config['directory']['main_file_dir'])
        self.sheet_1 = self.book.worksheets[0]
        self.max_rows_1 = self.sheet_1.max_row + 1

        self.treeview_inventar.delete(*self.treeview_inventar.get_children())
        self.treeview_inventar.grid_forget()

        self.treeview_inventar.tag_configure("oddrow", background="white")
        self.treeview_inventar.tag_configure("evenrow", background="gray95")

        for i in range(1, self.max_rows_1):
            for k in range(3, self.sheet_1.max_column + 1):
                if self.sheet_1.cell(row=i, column=k).value != None:
                    pass
                elif self.sheet_1.cell(row=i, column=k).value == None:
                    self.sheet_1.cell(row=i, column=k).value = ""

        for num, record in enumerate(range(1, self.max_rows_1)):
            if self.sheet_1.cell(row=record, column=1).value != None:
                if num % 2 != 0:
                    self.treeview_inventar.insert("", "end", iid=num, text="",
                                               values=(self.sheet_1.cell(row=record, column=1).value,
                                                       self.sheet_1.cell(row=record, column=2).value,
                                                       self.sheet_1.cell(row=record, column=3).value,
                                                       self.sheet_1.cell(row=record, column=4).value,
                                                       self.sheet_1.cell(row=record, column=5).value,
                                                       self.sheet_1.cell(row=record, column=6).value,
                                                       self.sheet_1.cell(row=record, column=7).value),
                                                       tags=("oddrow"))
                elif num % 2 == 0:
                    self.treeview_inventar.insert("", "end", iid=num, text="",
                                               values=(self.sheet_1.cell(row=record, column=1).value,
                                                       self.sheet_1.cell(row=record, column=2).value,
                                                       self.sheet_1.cell(row=record, column=3).value,
                                                       self.sheet_1.cell(row=record, column=4).value,
                                                       self.sheet_1.cell(row=record, column=5).value,
                                                       self.sheet_1.cell(row=record, column=6).value,
                                                       self.sheet_1.cell(row=record, column=7).value),
                                                       tags=("evenrow"))

        self.treeview_inventar.grid(row=0, column=0, sticky="nsew", pady=(35, 20), padx=40)
        self.sort_function("column1", self.treeview_inventar, False)

    def clicker_table_2(self, event):

        self.dialog_table2 = ctk.CTkToplevel(self)
        self.dialog_table2.geometry("260x290+1200+450")
        self.dialog_table2.resizable(False, False)
        self.dialog_table2.grab_set()
        self.dialog_table2.grid_columnconfigure(0, weight=1)
        self.dialog_table2.grid_columnconfigure(1, weight=1)

        self.artikel_table2_label = ctk.CTkLabel(self.dialog_table2, text="Artikel").grid(row=0, column=0, pady=(16, 4), sticky="e")
        self.hersteller_table2_label = ctk.CTkLabel(self.dialog_table2, text="Hersteller").grid(row=1, column=0, pady=4, sticky="e")
        self.model_table2_label = ctk.CTkLabel(self.dialog_table2, text="Model").grid(row=2, column=0, pady=4, sticky="e")
        self.sn_table2_label = ctk.CTkLabel(self.dialog_table2, text="Seriennummer").grid(row=3, column=0, pady=4, sticky="e")
        self.bemerkung_table2_label = ctk.CTkLabel(self.dialog_table2, text="Bemerkung").grid(row=4, column=0, pady=4, sticky="e")

        self.artikel_table2 = ctk.CTkEntry(self.dialog_table2)
        self.artikel_table2.grid(row=0, column=1, pady=(16, 4))
        self.hersteller_table2 = ctk.CTkEntry(self.dialog_table2)
        self.hersteller_table2.grid(row=1, column=1, pady=4)
        self.model_table2 = ctk.CTkEntry(self.dialog_table2)
        self.model_table2.grid(row=2, column=1, pady=4)
        self.sn_table2 = ctk.CTkEntry(self.dialog_table2)
        self.sn_table2.grid(row=3, column=1, pady=4)
        self.bemerkung_table2 = ctk.CTkEntry(self.dialog_table2)
        self.bemerkung_table2.grid(row=4, column=1, pady=4)

        self.selected_table2 = self.treeview_inventar.focus()
        self.values_table2 = self.treeview_inventar.item(self.selected_table2, 'values')

        self.dialog_table2.title(f"{self.values_table2[0]} {self.values_table2[1]}")

        self.artikel_table2.insert(0, self.values_table2[2])
        self.hersteller_table2.insert(0, self.values_table2[3])
        self.model_table2.insert(0, self.values_table2[4])
        self.sn_table2.insert(0, self.values_table2[5])
        self.bemerkung_table2.insert(0, self.values_table2[6])

        self.confirm_button_table2 = ctk.CTkButton(self.dialog_table2, text="OK",
                                                   command=self.update_record_table_2).grid(row=5, column=1, pady=(30, 4))
    def rueckgabe_fuction(self):

        rueckgabe_bestaetigen = messagebox.askyesno("Bitte bestätigen", "Sind Sie sicher?")

        if rueckgabe_bestaetigen:
            for rows in self.treeview_inventar.selection():
                for row_m in range(1,self.sheet_1.max_row):
                    introw = int(row_m) + 1

                    if self.sheet_1.cell(row=introw, column=3).value == self.treeview_inventar.item(rows, 'values')[2] \
                            and self.sheet_1.cell(row=introw, column=4).value == self.treeview_inventar.item(rows, 'values')[3] \
                            and self.sheet_1.cell(row=introw, column=5).value == self.treeview_inventar.item(rows, 'values')[4] \
                            and self.sheet_1.cell(row=introw, column=6).value == self.treeview_inventar.item(rows, 'values')[5]:
                        self.sheet_1.cell(row=introw, column=1).value = None
                        self.sheet_1.cell(row=introw, column=2).value = None
                        self.book.save(config['directory']['main_file_dir'])
                        break
            else:
                pass

        self.second_table_function()

    def update_record_table_2(self):
        self.treeview_inventar.item(self.selected_table2, text="",
                                    values=(self.values_table2[0], self.values_table2[1], self.artikel_table2.get(),
                                            self.hersteller_table2.get(),
                                            self.model_table2.get(),
                                            self.sn_table2.get(),
                                            self.bemerkung_table2.get()))

        for rows in range(1, self.max_rows_1):
            if self.sheet_1.cell(row=rows, column=3).value == self.values_table2[2] \
                    and self.sheet_1.cell(row=rows, column=4).value == self.values_table2[3] \
                    and self.sheet_1.cell(row=rows, column=5).value == self.values_table2[4] \
                    and self.sheet_1.cell(row=rows, column=6).value == self.values_table2[5] \
                    and self.sheet_1.cell(row=rows, column=7).value == self.values_table2[6]:
                self.sheet_1.cell(row=rows, column=3).value = self.artikel_table2.get()
                self.sheet_1.cell(row=rows, column=4).value = self.hersteller_table2.get()
                self.sheet_1.cell(row=rows, column=5).value = self.model_table2.get()
                self.sheet_1.cell(row=rows, column=6).value = self.sn_table2.get()
                self.sheet_1.cell(row=rows, column=7).value = self.bemerkung_table2.get()

        self.book.save(config['directory']['main_file_dir'])
        self.dialog_table2.destroy()
        if self.search.get() == "":
            self.second_table_function()
        else:
            self.search_function()

    def search_function(self):

        if self.search.get() == "":
            self.second_table_function()
        else:
            self.treeview_inventar.delete(*self.treeview_inventar.get_children())
            records = []
            for rows in range(1, self.max_rows_1):
                if self.sheet_1.cell(row=rows, column=1).value != None:
                    rec_sub = []
                    if self.search.get().lower() in self.sheet_1.cell(row=rows, column=1).value.lower():
                        for k in range(1,8):
                            rec_sub.append(self.sheet_1.cell(row=rows, column=k).value)
                        records.append(rec_sub)
                else:
                    pass

            self.treeview_inventar.tag_configure("oddrow", background="white")
            self.treeview_inventar.tag_configure("evenrow", background="gray95")

            count = 0
            for record in records:
                if count % 2 != 0:
                    self.treeview_inventar.insert("", "end", iid=count, text="", values=(
                        record[0], record[1], record[2], record[3], record[4], record[5], record[6]),
                                                  tags=("oddrow"))
                    count += 1
                elif count % 2 == 0:
                    self.treeview_inventar.insert("", "end", iid=count, text="", values=(
                        record[0], record[1], record[2], record[3], record[4], record[5], record[6]),
                                                  tags=("evenrow"))
                    count += 1

    def search_funktion_event(self, e):
        self.search_function()