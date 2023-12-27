import os
import subprocess
from datetime import date

import customtkinter as ctk
from PIL import Image
import openpyxl
import configparser

from docxtpl import DocxTemplate

config = configparser.ConfigParser()
config.read("config.ini", encoding='utf-8')
img_dir = config['directory']['images_dir']

date_today = date.today().strftime("%d.%m.%Y")

class FourthFrame(ctk.CTkFrame):

    def __init__(self, master):
        super().__init__(master, fg_color="transparent")

        self.grid_columnconfigure((0, 3), weight=1)
        self.grid_columnconfigure(1, minsize=145)
        self.grid_rowconfigure(0, minsize=145)

        self.users = ctk.CTkImage(Image.open(img_dir + "users.png"), size=(80, 80))
        self.word = ctk.CTkImage(Image.open(img_dir + "word.png"), size=(35, 35))

        # Label create
        self.vorname_label = ctk.CTkLabel(self, text="Vorname", font=ctk.CTkFont("Calibri", size=25, weight="bold"))
        self.vorname_label.grid(row=1, column=1, sticky="w", pady=15)
        self.nachname_label = ctk.CTkLabel(self, text="Nachname", font=ctk.CTkFont("Calibri", size=25, weight="bold"))
        self.nachname_label.grid(row=2, column=1, sticky="w", pady=15)
        self.abteilung_label = ctk.CTkLabel(self, text="Abteilung", font=ctk.CTkFont("Calibri", size=25, weight="bold"))
        self.abteilung_label.grid(row=3, column=1, sticky="w", pady=15)
        self.label_error = ctk.CTkLabel(self, text_color="red")

        # Entry create
        self.vorname_entry = ctk.CTkEntry(self, width=250, height=45, corner_radius=0, font=ctk.CTkFont("Calibri", size=23))
        self.vorname_entry.grid(row=1, column=2)
        self.nachname_entry = ctk.CTkEntry(self, width=250, height=45, corner_radius=0, font=ctk.CTkFont("Calibri", size=23))
        self.nachname_entry.grid(row=2, column=2)

        # Combo create
        list_abteilung = config['directory']['listen'].split(':')
        self.abteilung_combobox = ctk.CTkComboBox(self, values=list_abteilung, corner_radius=0, width=250,
                                                  height=45, font=ctk.CTkFont("Calibri", size=23),
                                                  dropdown_font=ctk.CTkFont("Calibri", size=23))
        self.abteilung_combobox.set(" Bitte auswählen")
        self.abteilung_combobox.grid(row=3, column=2)
        
        # Button create
        self.button_confirm = ctk.CTkButton(self, text="Hinzufügen", width=235, height=65,
                                            font=ctk.CTkFont("Calibri", size=28), corner_radius=0,
                                            command=lambda: self.mitarbeiter_add(
                                                            self.vorname_entry.get().strip().capitalize(),
                                                            self.nachname_entry.get().strip().capitalize(),
                                                            self.abteilung_combobox.get(), list_abteilung))
        self.button_confirm.grid(row=5, column=1, columnspan=2, pady=(65, 0))

        self.neue_user = False

        def checkbox_event():
            if check_var.get() == "on":
                self.neue_user = True
            else:
                self.neue_user = False

        check_var = ctk.StringVar(value="off")
        checkbox = ctk.CTkCheckBox(self, text="Neue Mitarbeiter", command=checkbox_event,
                                         font=ctk.CTkFont("Calibri", size=20),
                                         variable=check_var,
                                         onvalue="on", offvalue="off")
        checkbox.grid(row=4, column=2, pady=15, sticky="w")

    ### Functions ###

    def mitarbeiter_add(self, vorname, nachname, abteilung, list_abteilung):

        book = openpyxl.open(config['directory']['main_file_dir'])
        sheet = book.worksheets[1]
        max_rows = sheet.max_row + 1

        vorgesetzter_dict = {}
        for k, v in enumerate(list_abteilung):
            vorgesetzter_dict[list_abteilung[k]] = config['abteilungsleiter'][v]

        repeat_liste = []
        for i in range(1, max_rows):
            voll_name = str(sheet.cell(row=i, column=1).value.strip().lower()) + ' ' + str(sheet.cell(row=i, column=2).value.strip().lower())
            repeat_liste.append(voll_name)

        if vorname.lower() + ' ' + nachname.lower() in repeat_liste:
            self.label_error.configure(text=f"Der Mitarbeiter {vorname} {nachname}\nexistiert bereits", text_color="Yellow")
            self.label_error.grid(row=6, column=1, columnspan=2, sticky="n")
            self.after(4000, lambda: self.label_error.grid_forget())
            self.vorname_entry.delete(0, "end")
            self.nachname_entry.delete(0, "end")
            self.abteilung_combobox.set(" Bitte auswählen")
        elif len(vorname) == 0 or len(nachname) == 0 or abteilung == " Bitte auswählen":
            self.label_error.configure(text="Bitte füllen Sie alle Felder aus", text_color="red")
            self.label_error.grid(row=6, column=1, columnspan=2, sticky="n")
            self.after(4000, lambda: self.label_error.grid_forget())
        elif len(vorname) > 0 and len(nachname) > 0 and abteilung != "Bitte auswählen":
            sheet.cell(max_rows, column=1).value = vorname
            sheet.cell(max_rows, column=2).value = nachname
            sheet.cell(max_rows, column=3).value = abteilung
            sheet.cell(max_rows, column=4).value = vorgesetzter_dict[abteilung]

            try:
                book.save(config['directory']['main_file_dir'])

                four_frame_label_hinzu = ctk.CTkLabel(self, font=ctk.CTkFont("Calibri", size=22), text_color="#9fd8cb",
                                                      justify=ctk.LEFT, image=self.users, compound="top",
                                                      text=f"Mitarbeiter {vorname} {nachname}\nAbtelung:"
                                                           f" {abteilung}\nVorgesetzter: {vorgesetzter_dict[abteilung]}\n\nwurde "
                                                           f"erfolgreich hinzugefügt ✓", anchor="w")
                four_frame_label_hinzu.grid_forget()

                if self.neue_user:
                    self.neue_mitarbeiter(self.vorname_entry.get(), self.nachname_entry.get())
                else:
                    pass
                four_frame_label_hinzu.grid(row=6, column=1, columnspan=2, pady=(35, 0))
                self.after(4000, lambda: four_frame_label_hinzu.grid_forget())
                self.vorname_entry.delete(0, "end")
                self.nachname_entry.delete(0, "end")
                self.abteilung_combobox.set(" Bitte auswählen")
            except PermissionError:
                self.label_error.configure(text="Bitte schließen Sie die Excel-Datei", text_color="#F78154")
                self.label_error.grid(row=6, column=1, columnspan=2, sticky="n")
                self.after(4000, lambda: self.label_error.grid_forget())

    def neue_mitarbeiter(self, vorname, nachname):
        self.neue_mitarbeiter_dialog = ctk.CTkToplevel(self)
        self.neue_mitarbeiter_dialog.title(f"{vorname} {nachname}")
        self.neue_mitarbeiter_dialog.geometry(f"460x660+1200+450")
        self.neue_mitarbeiter_dialog.resizable(False, False)
        self.neue_mitarbeiter_dialog.grab_set()
        self.neue_mitarbeiter_dialog.grid_columnconfigure(0, weight=1)
        self.neue_mitarbeiter_dialog.grid_columnconfigure(1, weight=1)

        self.geschlecht = ""

        def radiobutton_event():
            if radio_var.get() == 1:
                self.geschlecht = "Herr"
            else:
                self.geschlecht = "Frau"

        radio_var = ctk.IntVar(0)
        self.geschlecht_mann = ctk.CTkRadioButton(self.neue_mitarbeiter_dialog, text="Mann",
                                                  command=radiobutton_event,
                                                  font=ctk.CTkFont("Calibri", size=20),
                                                  variable= radio_var, value=1)
        self.geschlecht_frau = ctk.CTkRadioButton(self.neue_mitarbeiter_dialog,
                                                  font=ctk.CTkFont("Calibri", size=20), text="Frau",
                                                  command=radiobutton_event, variable= radio_var, value=2)
        self.geschlecht_mann.grid(row=0, column=0, pady=20)
        self.geschlecht_frau.grid(row=0, column=1, pady=20)

        self.nm_vorname_label = ctk.CTkLabel(self.neue_mitarbeiter_dialog, text="Vorname",
                                             font=ctk.CTkFont("Calibri", size=20)).grid(row=1, column=0, padx=(50,0), pady=20, sticky="w")
        self.nm_nachname_label = ctk.CTkLabel(self.neue_mitarbeiter_dialog, text="Nachname",
                                             font=ctk.CTkFont("Calibri", size=20)).grid(row=2, column=0, padx=(50,0), pady=20, sticky="w")
        self.nm_datum_label = ctk.CTkLabel(self.neue_mitarbeiter_dialog, text="Datum",
                                             font=ctk.CTkFont("Calibri", size=20)).grid(row=3, column=0, padx=(50,0), pady=20, sticky="w")
        self.nm_windows_ps = ctk.CTkLabel(self.neue_mitarbeiter_dialog, text="Windows Password",
                                             font=ctk.CTkFont("Calibri", size=20)).grid(row=4, column=0, padx=(50,0), pady=20, sticky="w")
        self.nm_quorra_ps = ctk.CTkLabel(self.neue_mitarbeiter_dialog, text="Quorra Password",
                                             font=ctk.CTkFont("Calibri", size=20)).grid(row=5, column=0, padx=(50,0), pady=20, sticky="w")

        self.nm_vorname_entry = ctk.CTkEntry(self.neue_mitarbeiter_dialog,
                                             font=ctk.CTkFont("Calibri", size=20, weight="bold"),
                                             corner_radius=0,
                                             width=150,
                                             height=35)
        self.nm_vorname_entry.insert("0", vorname)

        self.nm_nachname_entry = ctk.CTkEntry(self.neue_mitarbeiter_dialog,
                                             font=ctk.CTkFont("Calibri", size=20, weight="bold"),
                                             corner_radius=0,
                                             width=150,
                                             height=35)
        self.nm_nachname_entry.insert("0", nachname)

        self.nm_datum_entry = ctk.CTkEntry(self.neue_mitarbeiter_dialog,
                                             font=ctk.CTkFont("Calibri", size=20, weight="bold"),
                                             corner_radius=0,
                                             width=150,
                                             height=35)
        self.nm_datum_entry.insert("0", date_today)
        self.nm_windows_entry = ctk.CTkEntry(self.neue_mitarbeiter_dialog,
                                             font=ctk.CTkFont("Calibri", size=20, weight="bold"),
                                             corner_radius=0,
                                             width=150,
                                             height=35)
        self.nm_quorra_entry = ctk.CTkEntry(self.neue_mitarbeiter_dialog,
                                             font=ctk.CTkFont("Calibri", size=20, weight="bold"),
                                             corner_radius=0,
                                             width=150,
                                             height=35)

        self.nm_vorname_entry.grid(row=1, column=1, pady=20)
        self.nm_nachname_entry.grid(row=2, column=1, pady=20)
        self.nm_datum_entry.grid(row=3, column=1, pady=20)
        self.nm_windows_entry.grid(row=4, column=1, pady=20)
        self.nm_quorra_entry.grid(row=5, column=1, pady=20)

        self.nm_bestaetigung_button = ctk.CTkButton(self.neue_mitarbeiter_dialog,
                                                    text="Daten erstellen",
                                                    corner_radius=0,
                                                    width=190,
                                                    height=40,
                                                    hover_color="#5FAD56",
                                                    font=ctk.CTkFont("Calibri", size=20, weight="bold"),
                                                    command=lambda : self.neue_mitarbeiter_word(self.geschlecht,
                                                                                                self.nm_vorname_entry.get(),
                                                                                                self.nm_nachname_entry.get(),
                                                                                                self.nm_datum_entry.get(),
                                                                                                self.nm_windows_entry.get(),
                                                                                                self.nm_quorra_entry.get())).grid(row=6,
                                                                                                column=0,
                                                                                                columnspan=2,
                                                                                                pady=(50,20))

    def neue_mitarbeiter_word(self, geschlecht, vorname, nachname, datum, wpassword, qpassword):

        if geschlecht=="" or len(vorname)<1 or len(nachname)<1 or len(datum)<1 or len(wpassword)<1:
            pass
        else:

            contex = {'geschlecht': geschlecht,
                      'vorname': vorname,
                      'nachname': nachname,
                      'datum': datum,
                      'wpassword': wpassword,
                      'qpassword': qpassword}

            mitarbeiter_directory = config['directory']['mitarbeiter_directory']

            file_dir = (config['directory']['mitarbeiter_default'])
            doc = DocxTemplate(file_dir)
            doc.render(contex)

            vollname_string = f"{vorname.strip()}_{nachname}"
            folder_path = (mitarbeiter_directory + vollname_string+ "//")

            if not os.path.exists(folder_path):
                os.makedirs(folder_path)

            dir_path = (mitarbeiter_directory + vollname_string + f"//Neuer_Mitarbeiter_{vollname_string}.docx")
            doc.save(dir_path)

            def open_nm(path):
                subprocess.Popen(['start', path], shell=True)

            self.word_open = ctk.CTkButton(self.neue_mitarbeiter_dialog,
                                           text="Daten öffnen",
                                           corner_radius=0,
                                           image=self.word,
                                           width=190,
                                           height=40,
                                           font=ctk.CTkFont("Calibri", size=20, weight="bold"),
                                           command=lambda: open_nm(dir_path)).grid(row=7, column=0, columnspan=2, pady=20)